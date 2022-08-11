# -*- coding: utf-8 -*-
"""
Created on Sun Jul 19 18:20:11 2020
Copyright 2022 Guoqi Liu

   Licensed under the Apache License, Version 2.0 (the "License");
   you may not use this file except in compliance with the License.
   You may obtain a copy of the License at

     http://www.apache.org/licenses/LICENSE-2.0

   Unless required by applicable law or agreed to in writing, software
   distributed under the License is distributed on an "AS IS" BASIS,
   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
   See the License for the specific language governing permissions and
   limitations under the License.
@author: Guoqi Liu  642847452@qq.com  
"""


import tkinter as tk
from tkinter import *
from tkinter import ttk
import tkinter.messagebox
import tkinter.simpledialog


from tkinter.filedialog import askdirectory
import matplotlib.pyplot as plt
import warnings
from PIL import Image,ImageTk
import logging
#import wmi
import base64
from pyDes import *
import sys
import re
import csv
import os
import os.path
import json
import math
from math import *
import time
import pandas as pd
import numpy as np
import xlwt,xlrd
from scipy.optimize import curve_fit
from openpyxl import load_workbook

#warnings.filterwarnings('ignore')

LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
DATE_FORMAT = "%m/%d/%Y %H:%M:%S %p"

logging.basicConfig(filename='my.log', level=logging.INFO, format=LOG_FORMAT, datefmt=DATE_FORMAT)

def Age76Pb(Rap76):
    #Rap76=0.058
    
    Tmin = 0.001
    Tmax = 4556
    N = 0
    if Rap76 > 0.0460455:
        Tm = (Tmax + Tmin) / 2
        Rap = (exp(0.00098485 * Tm) - 1) / (exp(0.000155125 * Tm) - 1) / 137.88      
    
        
        while abs(Rap76 - Rap)>0.00005 and N <10:
            delta = Rap76 - Rap
            if Rap < Rap76:
                Tmin = Tm
            else:Tmax = Tm
            Rapi = (exp(0.00098485 * Tmin) - 1) / (exp(0.000155125 * Tmin) - 1) / 137.88
            Raps = (exp(0.00098485 * Tmax) - 1) / (exp(0.000155125 * Tmax) - 1) / 137.88
            Tm = Tmin + (Tmax - Tmin) * (Rap76 - Rapi) / (Raps - Rapi)
            Age76Pb = Tm
            
            N = N + 1
    else:Age76Pb = 0
    return Age76Pb

def regression(xv,num206):
    def func(xv, a, b,c):    
        if regression_method==0: 
            c=0
            return a*xv+b
        if regression_method==1:
            c=0
            return a*np.log(xv)+b
        if regression_method==2:
            return a*np.sqrt(xv)*(b*np.square(xv)+c)
        if regression_method==3:
            c=0
            return a*(xv**b)+c
        else:
            pass
    
    #xv=np.arange(len(num206))
    stdname=1
    
    Mun_std=1
    y206=np.array(num206)
    popt206, pcov206 = curve_fit(func, xv, y206)
    a206 = popt206[0] 
    b206 = popt206[1]
    c206 = popt206[2]
    yvals206 = func(xv,a206,b206,c206)
  
        
    
    
    return yvals206


    
def loaddata(name,isoname):
    
    if ele.get()==0:
        
        
        data_all=np.loadtxt(inputpath+'//'+name,dtype=str,delimiter=',',skiprows=13,comments='#')
        col_data_name=data_all[0,:].tolist()
        

        x=data_all[2:,col_data_name.index('Time')].astype(float)
        try:
            y2=data_all[2:,col_data_name.index('204Pb')].astype(float)
        except:
            y2=x*0
        try:
            y1=data_all[2:,col_data_name.index('202Hg')].astype(float)
            
        except(ValueError):
            #logging.warning('202Hg Not found!')
            
            y1=x*0
        
        y3=data_all[2:,col_data_name.index('206Pb')].astype(float)
        y4=data_all[2:,col_data_name.index('207Pb')].astype(float)
        y5=data_all[2:,col_data_name.index('208Pb')].astype(float)
        y6=data_all[2:,col_data_name.index('232Th')].astype(float)
        y7=data_all[2:,col_data_name.index('238U')].astype(float)
        return x,y1,y2,y3,y4,y5,y6,y7
    elif ele.get()==1:
        #print(name)
        nameD=name.replace('.csv','.D')
        inputpath_sec=os.path.join(inputpath,nameD)
        #print(inputpath_sec)
  
        try:
            
            logging.info('Agilent csv :date begin from the third row.')
            data_all=np.loadtxt(inputpath_sec+'//'+name,dtype=str,delimiter=',',skiprows=3,comments='          ')

            col_data_name=data_all[0,:].tolist()
            
            col_name=[]
            for i in range(len(col_data_name)):
                col_name.append(''.join(list(filter(str.isdigit, col_data_name[i]))))
            
                
               
            x=data_all[2:,0].astype(float)
            try:
                y1=data_all[2:,col_data_name.index(isoname[1])].astype(float)
                
            except(ValueError):                
                logging.warning('202Hg Not found!')
                y1=x*0
            try:
               
               y2=data_all[2:,col_name.index(isoname[2])].astype(float)
            except(ValueError): 
                y2=x*0
            #y2=data_all[2:,col_name.index(isoname[2])].astype(float)
            y3=data_all[2:,col_name.index(isoname[3])].astype(float)
            y4=data_all[2:,col_name.index(isoname[4])].astype(float)
            y5=data_all[2:,col_name.index(isoname[5])].astype(float)
            y6=data_all[2:,col_name.index(isoname[6])].astype(float)
            y7=data_all[2:,col_name.index(isoname[7])].astype(float)
            return x,y1,y2,y3,y4,y5,y6,y7              
                
                
            
  
 
        except ValueError:
            logging.info('Agilent csv :date begin from the Sencond row.')
            data_all=np.loadtxt(inputpath_sec+'//'+name,dtype=str,delimiter=',',skiprows=2,comments='          ')
            col_data_name=data_all[0,:].tolist()
            col_name=[]
            for i in range(len(col_data_name)):
                col_name.append(''.join(list(filter(str.isdigit, col_data_name[i]))))
            
            x=data_all[2:,0].astype(float)
            
            try:
                y1=data_all[2:,col_data_name.index(isoname[1])].astype(float)
            except(ValueError):
                #logging.warning('202Hg Not found!')
                y1=x*0
            try:
               
               y2=data_all[2:,col_name.index(isoname[2])].astype(float)
            except(ValueError): 
                y2=x*0
            #y2=data_all[2:,col_name.index(isoname[2])].astype(float)
            y3=data_all[2:,col_name.index(isoname[3])].astype(float)
            y4=data_all[2:,col_name.index(isoname[4])].astype(float)
            y5=data_all[2:,col_name.index(isoname[5])].astype(float)
            try:
                y6=data_all[2:,col_name.index(isoname[6])].astype(float)
            except(ValueError):
                logging.warning('232Th Not found!')
                y6=x*0
                
            y7=data_all[2:,col_name.index(isoname[7])].astype(float)
            return x,y1,y2,y3,y4,y5,y6,y7
            
            
       
            
    elif ele.get()==2:
        
        
        data_all=np.loadtxt(inputpath+'//'+name,dtype=str,delimiter=',',skiprows=7,comments='#')
        col_data_name=data_all[0,:].tolist()
        
        try:
            
            x=data_all[2:,col_data_name.index(isoname[0])].astype(float)
            try:
                y1=data_all[2:,col_data_name.index(isoname[1])].astype(float)
            except (ValueError):
                y1=x*0
                #logging.warning('202Hg Not found!')
            y2=data_all[2:,col_data_name.index(isoname[2])].astype(float)
            y3=data_all[2:,col_data_name.index(isoname[3])].astype(float)
            y4=data_all[2:,col_data_name.index(isoname[4])].astype(float)
            y5=data_all[2:,col_data_name.index(isoname[5])].astype(float)
            y6=data_all[2:,col_data_name.index(isoname[6])].astype(float)
            y7=data_all[2:,col_data_name.index(isoname[7])].astype(float)
            return x,y1,y2,y3,y4,y5,y6,y7
        
        except Exception as e:
            logging.error('Unknown erro:%s',str(e))

            print(str(e))
    

 
            
def Age_Calculate_average():
    logging.info("Average calculation method")
    global NIST_STD
    M1=tk.messagebox.askyesno(title = 'Are trace elements calculated?',message="Is it needed for elemental content calculation?")
    if M1:
        
        NIST_STD=tk.simpledialog.askstring(title = 'Name of Standard',prompt='Name of Standard',initialvalue = 'NIST610')

    else:
        
        pass
    try:
        
        name=np.loadtxt(outputpath+'//'+'result_all.csv',dtype=str,delimiter=',',skiprows=1,usecols=(2))
        for i in range(len(name)):
            name[i]=name[i].strip()
            
        File_name=np.loadtxt(outputpath+'//'+'result_all.csv',dtype=str,delimiter=',',skiprows=1,usecols=(1))
        date=np.loadtxt(outputpath+'//'+'result_all.csv',delimiter=',',skiprows=1,usecols=(3,4,5,6,7,8,9,10,11,12,13,14,15,16,18,19,20,21,22))
        print(Standard_names)
      
        date_all=pd.DataFrame(date,index=name).astype(float)
      
        def coefficient(NIST_STD):
            try:
                if NIST_STD in ['NIST610','NIST 610','SRM610','SRM 610']:
                    logging.info('NIST 610 as  external standard!')
                    coefficient_U=461.5/date_all[14][NIST_STD][:].mean()
                    coefficient_Th=457.2/date_all[15][NIST_STD][:].mean()
                    coefficient_Pb=426/(date_all[16][NIST_STD][:].mean()+date_all[17][NIST_STD][:].mean()+date_all[18][NIST_STD][:].mean())                      
                elif NIST_STD  in ['NIST612','NIST 612','SRM612','SRM 612']:
                    logging.info('NIST 612 as  external standard!')
                    coefficient_U=37.38/date_all[14][NIST_STD][:].mean()
                    coefficient_Th=37.79/date_all[15][NIST_STD][:].mean()
                    coefficient_Pb=38.57/(date_all[16][NIST_STD][:].mean()+date_all[17][NIST_STD][:].mean()+date_all[18][NIST_STD][:].mean())        
                elif NIST_STD in ['NIST614','NIST 614','SRM614','SRM 614']:
                    logging.info('NIST 614 as  external standard!')
                    coefficient_U=0.832/date_all[14][NIST_STD][:].mean()
                    coefficient_Th=0.748/date_all[15][NIST_STD][:].mean()
                    coefficient_Pb=2.32/(date_all[16][NIST_STD][:].mean()+date_all[17][NIST_STD][:].mean()+date_all[18][NIST_STD][:].mean())
                else:
                    coefficient_U=0
                    coefficient_Th=0
                    coefficient_Pb=0                
            except Exception as e:
                logging.error('Unknown erro:%s',str(e))
                print(str(e))
                coefficient_U=0
                coefficient_Th=0
                coefficient_Pb=0
            return coefficient_U,coefficient_Th,coefficient_Pb
        coefficient_U,coefficient_Th,coefficient_Pb=coefficient(NIST_STD)

       
        
        def factor(n,stand_v,i):
            if n%2==0:
                if n<=2:
                    f=stand_v/date_all[i][standard][0:2].mean()
                    return f
                f=(factor(n-2,stand_v,i)+factor(n-4,stand_v,i))*0.5
                return f
            else :
                n=n-1
                if n<=2:
                    f=stand_v/date_all[i][standard][0:2].mean()
                    return f
                f=(factor(n-2,stand_v,i)+factor(n-4,stand_v,i))*0.5
                return f
            

        
            

                
            
            
        try :
            num_data=len(date_all[0][standard])
        except KeyError:
            print('The standard NAME is incorrect！')
            tk.messagebox.showinfo(title='Information！', message='Please check the standard NAME！')
            logging.error('Wrong standard NAME!')
            
        f207_206=factor(num_data,P382,0)
        
        f206_238=factor(num_data,a,2)
        

        f207_235=factor(num_data,b,4)
        

        f208_232=factor(num_data,c,6)
        
        f208_206=1
        f232_206=1/f208_232
        f208_204=1
       
        
        result_cal_207Pb_206Pb=date_all[0]*f207_206
        
        result_cal_206Pb_238U=date_all[2]*f206_238
        result_cal_207Pb_235U=date_all[4]*f207_235
        result_cal_208Pb_232Th=date_all[6]*f208_232
        result_cal_208Pb_206Pb=date_all[8]*f208_206
        result_cal_232Th_206Pb=date_all[10]*f232_206
        result_cal_208Pb_204Pb=date_all[12]*f208_204



        try:
            result_cal_age=[]
            for i in range(len(date_all)):
                w=(sqrt(pow((date_all[3][i]/result_cal_206Pb_238U[i]),2)+0.0009)*result_cal_206Pb_238U[i])/result_cal_206Pb_238U[i]*100
                s=(sqrt(pow((date_all[5][i]/result_cal_207Pb_235U[i]),2)+0.0009)*result_cal_207Pb_235U[i])/result_cal_207Pb_235U[i]*100
                o=(sqrt(pow((date_all[1][i]/result_cal_207Pb_206Pb[i]),2)+0.0001)*result_cal_207Pb_206Pb[i])/result_cal_207Pb_206Pb[i]*100
                if (abs((w*w+s*s-o*o)/(2*w*s)))<w/s:
                    rho=abs((w*w+s*s-o*o)/(2*w*s))
                else :
                    rho=w/s
                s002=[File_name[i],name[i],result_cal_207Pb_206Pb[i],date_all[1][i],date_all[2][i]*f206_238,date_all[3][i],result_cal_207Pb_235U[i],date_all[5][i],result_cal_208Pb_232Th[i],date_all[7][i],\
                      result_cal_208Pb_206Pb[i],date_all[9][i],
                      result_cal_232Th_206Pb[i],date_all[11][i],result_cal_208Pb_204Pb[i],date_all[13][i],
                      '-------------',log(abs(date_all[2][i]*f206_238+1))/(0.000000000155125)/1000000,\
                      (sqrt(pow(date_all[3][i]/(date_all[2][i]*f206_238),2)+0.0009))*(log(abs(date_all[2][i]*f206_238+1))/(0.000000000155125)/1000000),\
                      log(abs(result_cal_207Pb_235U[i]+1))/(0.00000000098485)/1000000,\
                      (sqrt(pow(date_all[5][i]/result_cal_207Pb_235U[i],2)+0.0009))*(log(abs(result_cal_207Pb_235U[i]+1))/(0.00000000098485)/1000000),\
                      log(abs(result_cal_208Pb_232Th[i]+1))/(0.000000000049475)/1000000,\
                      (sqrt(pow(date_all[7][i]/result_cal_208Pb_232Th[i],2)+0.0009))*(log(abs(result_cal_208Pb_232Th[i]+1))/(0.000000000049475)/1000000),\
                       Age76Pb(result_cal_207Pb_206Pb[i]),Age76Pb(result_cal_207Pb_206Pb[i]+date_all[1][i])-Age76Pb(result_cal_207Pb_206Pb[i]),\
                      '-------------',(1/(date_all[2][i]*f206_238)),(1/(date_all[2][i]*f206_238))*sqrt(pow(date_all[3][i]/(date_all[2][i]*f206_238),2)+0.0009),\
                          result_cal_207Pb_206Pb[i],date_all[1][i],'-------------',result_cal_207Pb_235U[i],\
                      (sqrt(pow(date_all[5][i]/result_cal_207Pb_235U[i],2)+0.0009))*result_cal_207Pb_235U[i],\
                      result_cal_206Pb_238U[i],sqrt(pow((date_all[3][i]/result_cal_206Pb_238U[i]),2)+0.0009)*result_cal_206Pb_238U[i],rho,\
                          '-------------',result_cal_208Pb_232Th[i],\
                      sqrt(pow(date_all[7][i]/result_cal_208Pb_232Th[i],2)+0.0009)*result_cal_208Pb_232Th[i],\
                      result_cal_208Pb_206Pb[i],date_all[9][i],'-------------',abs(date_all[14][i]*coefficient_U),\
                      date_all[15][i]*coefficient_Th,(date_all[16][i]+date_all[17][i]+date_all[18][i])*coefficient_Pb]

                
                
                
                result_cal_age.append(s002)


        except Exception as e:
            logging.error('Unknown erro:%s',str(e))
            print(str(e))
            tk.messagebox.showinfo(title='Information！', message='Standard Name not set or“result_all”file error！')
            print('Unknown erro！')
            


        
        #print(result_cal_age)
        
        
        
        

            
        
        os.chdir(outputpath)

        workbook = xlwt.Workbook()
        worksheet=workbook.add_sheet('My Worksheet',cell_overwrite_ok=True)
        style = xlwt.XFStyle() 
        font = xlwt.Font() 
        font.name = 'Times New Roman' 
        font.bold = True 
        font.underline = True 
        font.italic = True 
        style.font = font 
        
        worksheet.write(0, 0, 'FileNames')
        worksheet.write(0, 1, 'SampleName')
        worksheet.write(0, 2, '207Pb/206Pb')
        worksheet.write(0, 3, '2s')
        worksheet.write(0, 4, '206Pb/238U')
        worksheet.write(0, 5, '2s')

        worksheet.write(0, 6, '207Pb/235U')
        worksheet.write(0, 7, '2s')

        worksheet.write(0, 8, '208Pb/232Th')
        worksheet.write(0, 9, '2s')
        worksheet.write(0, 10, '208Pb/206Pb')
        worksheet.write(0, 11, '2s')
        
        worksheet.write(0, 12, '232Th/206Pb')
        worksheet.write(0, 13, '2s')
        worksheet.write(0, 14, '208Pb/204Pb')
        worksheet.write(0, 15, '2s')
        
        worksheet.write(0, 16, 'Age')
        worksheet.write(0, 17, '206Pb_238U Age(Ma)')
        worksheet.write(0, 18, '2s')
        worksheet.write(0, 19, '207Pb_235U Age(Ma)')
        worksheet.write(0, 20, '2s')    
        worksheet.write(0, 21, '208Pb_232Th  Age(Ma)')
        worksheet.write(0, 22, '2s')
        worksheet.write(0, 23, '207Pb_206Pb  Age(Ma)')
        worksheet.write(0, 24, '2s')

        worksheet.write(0, 25, 'Terra-Wasserburg Plot:')
        worksheet.write(0, 26, '238U/206Pb')
        worksheet.write(0, 27, '2s error')
        worksheet.write(0, 28, '207Pb/206Pb')
        worksheet.write(0, 29, '2s error')
        
        worksheet.write(0, 30, 'Plotting purposes:')
        
        worksheet.write(0, 31, '207Pb/235U')
        worksheet.write(0, 32, '2s error')
        worksheet.write(0, 33, '206Pb/238U')
        worksheet.write(0, 34, '2s error')
        worksheet.write(0, 35, 'rho')
        worksheet.write(0, 36, ' ')
        
        worksheet.write(0, 37, '208Pb/232Th')
        worksheet.write(0, 38, '2s error')
        worksheet.write(0, 39, '208Pb/206Pb')
        worksheet.write(0, 40, '2s error')       
        worksheet.write(0, 41, 'Trace element')
        worksheet.write(0, 42, 'U(ppm)')
        worksheet.write(0, 43, 'Th(ppm)')
        worksheet.write(0, 44, 'Pb(ppm)')  
        

        
        
        
        for row in range(len(result_cal_age)):
            for col in range(45):
                worksheet.write(row+1,col,result_cal_age[row][col])
            
            
        
        now = time.strftime("%H%M%S",time.localtime(time.time()))     
        workbook.save('cal_age_result_A'+now+'.xls')
        
        print('The calculation was successful.The file saved in：',outputpath)
        logging.info('The calculation was successful.The file saved in:%s',outputpath)
    except ValueError:
        tk.messagebox.showinfo(title='Information！', message='“result_all.csv"file not found or  the format is incorrect！')
        
        logging.critical('“result_all.csv" were not found or the format is incorrect.')    
    except NameError:
        tk.messagebox.showinfo(title='Information！', message='Output directory or standard setting error！')
       
        logging.error('The output directory is not set.') 
    except OSError:
        tk.messagebox.showinfo(title='Information！', message='"result_all.csv" not found！')
        logging.error('"result_all.csv" not found')
        print('"result_all.csv" not found')
    
    tk.messagebox.showinfo(title='Congratulations！', message='The age calculation was successful！')
        
def Age_Calculate():
    logging.info("Linear calculation method")
   
    try:
        name=np.loadtxt(outputpath+'//'+'result_all.csv',dtype=str,delimiter=',',skiprows=1,usecols=(2))
        Fname=np.loadtxt(outputpath+'//'+'result_all.csv',dtype=str,delimiter=',',skiprows=1,usecols=(1))
        File_name=Fname.tolist()
        ssf=name.tolist()
        date=np.loadtxt(outputpath+'//'+'result_all.csv',delimiter=',',skiprows=1,usecols=(3,4,5,6,7,8,9,10,11,12,13,14,15,16))
        
        date_all=pd.DataFrame(date,index=name).astype(float)     
        
        
        myix=[i for i,x in enumerate(ssf) if x==standard]
        myiy=[i for i,x in enumerate(ssf) if x!=standard]       
            
        sample_slected_pd=pd.DataFrame(columns=['0','1','2','3','4','5','6','7','8','9','10','11','12','13'])

        for i in range(len(myix)):
            
            if i==0:
                if myix==0:
                    continue
                else:
                   sample_slected_pd=sample_slected_pd.append(date_all[:][:myix[0]], ignore_index=True) 
            else:
                
                if myix[i]-myix[i-1]==1:
                    continue
                sample_slected_pd=sample_slected_pd.append(date_all[:][myix[i-1]+1:myix[i]], ignore_index=True)
        
                if i==len(myix)-1:
                    sample_slected_pd=sample_slected_pd.append(date_all[:][myix[i]+1:], ignore_index=True)
        
        def func(lst):
            new_list = []
            for i,j in zip(lst,lst[1:]):
                    if j - i > 1:
                            new_list.append(lst[:lst.index(j)])
                            lst = lst[lst.index(j):]
            new_list.append(lst)
            return new_list

        class_sample=func(myiy)
        class_std=func(myix)
        date_all_num=pd.DataFrame(date).astype(float)
        
        
            
        def factor(SEP,stand_v,i,class_std,class_sample):
            
            if len(class_std)==len(class_sample):
                if list(class_std[0])[0]==0:
                    class_std.append(class_std[-1])
                    f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                    return f
                else:
                     class_std.insert(0,class_std[0])
                     f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                     return f
            elif len(class_std)<len(class_sample):
                class_std.append(class_std[-1])
                class_std.insert(0,class_std[0])
                
                f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                return f
            else:
                f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                return f
            
               
                
        result_cal_age=[]
        try:
            
            
            for SEP,each in enumerate(class_sample):
                
                 
                n=len(each)
                w_1=1
                
                f207_206A=factor(SEP,P382,0,class_std,class_sample)    
                f206_238A=factor(SEP,a,2,class_std,class_sample) 
                f207_235A=factor(SEP,b,4,class_std,class_sample)  
                f208_232A=factor(SEP,c,6,class_std,class_sample)    
                f208_206A=1
                f232_206A=1/f208_232A
                f208_204A=1
                f207_206B=factor(SEP+1,P382,0,class_std,class_sample)    
                f206_238B=factor(SEP+1,a,2,class_std,class_sample) 
                f207_235B=factor(SEP+1,b,4,class_std,class_sample)  
                f208_232B=factor(SEP+1,c,6,class_std,class_sample)    
                f208_206B=1
                f232_206B=1/f208_232B
                f208_204B=1         
               
                for i in each:
                    
                    result_cal_207Pb_206Pb=date_all[0][i]*(w_1/n)*f207_206B+date_all[0][i]*(1-(w_1/n))*f207_206A
            
                    result_cal_206Pb_238U=date_all[2][i]*(w_1/n)*f206_238B+date_all[2][i]*(1-(w_1/n))*f206_238A
                    result_cal_207Pb_235U=date_all[4][i]*(w_1/n)*f207_235B+date_all[4][i]*(1-(w_1/n))*f207_235A
                    result_cal_208Pb_232Th=date_all[6][i]*(w_1/n)*f208_232B+date_all[6][i]*(1-(w_1/n))*f208_232A
                    result_cal_208Pb_206Pb=date_all[8][i]*(w_1/n)*f208_206B+date_all[8][i]*(1-(w_1/n))*f208_206A
                    result_cal_232Th_206Pb=date_all[10][i]*(w_1/n)*f232_206B+date_all[10][i]*(1-(w_1/n))*f232_206A
                    result_cal_208Pb_204Pb=date_all[12][i]*(w_1/n)*f208_204B+date_all[12][i]*(1-(w_1/n))*f208_204A
                    
                    w=(sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U)/result_cal_206Pb_238U*100
                    s=(sqrt(pow((date_all[5][i]/result_cal_207Pb_235U),2)+0.0009)*result_cal_207Pb_235U)/result_cal_207Pb_235U*100
                    o=(sqrt(pow((date_all[1][i]/result_cal_207Pb_206Pb),2)+0.0001)*result_cal_207Pb_206Pb)/result_cal_207Pb_206Pb*100
                    if (abs((w*w+s*s-o*o)/(2*w*s)))<w/s:
                        rho=abs((w*w+s*s-o*o)/(2*w*s))
                    else :
                        rho=w/s
                    s002=[File_name[i],ssf[i],result_cal_207Pb_206Pb,date_all[1][i],result_cal_206Pb_238U,date_all[3][i],result_cal_207Pb_235U,date_all[5][i],result_cal_208Pb_232Th,date_all[7][i],\
                          result_cal_208Pb_206Pb,date_all[9][i],
                          result_cal_232Th_206Pb,date_all[11][i],result_cal_208Pb_204Pb,date_all[13][i],
                          '-------------',log(abs(result_cal_206Pb_238U+1))/(0.000000000155125)/1000000,(sqrt(pow(date_all[3][i]/result_cal_206Pb_238U,2)+0.0009))*(log(abs(result_cal_206Pb_238U+1))/(0.000000000155125)/1000000),\
                          log(abs(result_cal_207Pb_235U+1))/(0.00000000098485)/1000000,(sqrt(pow(date_all[5][i]/result_cal_207Pb_235U,2)+0.0009))*(log(abs(result_cal_207Pb_235U+1))/(0.00000000098485)/1000000),\
                          log(abs(result_cal_208Pb_232Th+1))/(0.000000000049475)/1000000,(sqrt(pow(date_all[7][i]/result_cal_208Pb_232Th,2)+0.0009))*(log(abs(result_cal_208Pb_232Th+1))/(0.000000000049475)/1000000),\
                          Age76Pb(result_cal_207Pb_206Pb),Age76Pb(result_cal_207Pb_206Pb+date_all[1][i])-Age76Pb(result_cal_207Pb_206Pb),'------------',1/result_cal_206Pb_238U,(1/result_cal_206Pb_238U)*(sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U)/result_cal_206Pb_238U,\
                          result_cal_207Pb_206Pb,date_all[1][i],'------------',result_cal_207Pb_235U,result_cal_207Pb_235U*sqrt(pow(date_all[5][i]/result_cal_207Pb_235U,2)+0.0009),result_cal_206Pb_238U,sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U,\
                          rho,                       
                             
                          result_cal_208Pb_232Th,date_all[7][i],result_cal_208Pb_206Pb,date_all[9][i]]
                    
                    
                    result_cal_age.append(s002)
                    w_1=w_1+1
        except Exception as e:
            print(str(e))
            logging.error('Unknown erro:%s',str(e))
            tk.messagebox.showinfo(title='information！', message='Standard not set or “result_all”file is incorrect！')
           
                
                
                
        

            
        
        os.chdir(outputpath)

        workbook = xlwt.Workbook()
        worksheet=workbook.add_sheet('My Worksheet',cell_overwrite_ok=True)
        style = xlwt.XFStyle() 
        font = xlwt.Font()
        font.name = 'Times New Roman' 
        font.bold = True 
        font.underline = True 
        font.italic = True 
        style.font = font 
        worksheet.write(0, 0, 'File Names')
        worksheet.write(0, 1, 'Sample Names')
        worksheet.write(0, 2, '207Pb/206Pb')
        worksheet.write(0, 3, '2s')
        worksheet.write(0, 4, '206Pb/238U')
        worksheet.write(0, 5, '2s')

        worksheet.write(0, 6, '207Pb/235U')
        worksheet.write(0, 7, '2s')

        worksheet.write(0, 8, '208Pb/232Th')
        worksheet.write(0, 9, '2s')
        worksheet.write(0, 10, '208Pb/206Pb')
        worksheet.write(0, 11, '2s')
        
        worksheet.write(0, 12, '232Th/206Pb')
        worksheet.write(0, 13, '2s')
        worksheet.write(0, 14, '208Pb/204Pb')
        worksheet.write(0, 15, '2s')
        
        worksheet.write(0, 16, 'Age')
        worksheet.write(0, 17, '206Pb_238U Age(Ma)')
        worksheet.write(0, 18, '2s')
        worksheet.write(0, 19, '207Pb_235U Age(Ma)')
        worksheet.write(0, 20, '2s')    
        worksheet.write(0, 21, '208Pb_232Th  Age(Ma)')
        worksheet.write(0, 22, '2s')
        
        worksheet.write(0, 23, '207Pb_206Pb  Age(Ma)')
        worksheet.write(0, 24, '2s')

        
        worksheet.write(0, 25, 'Terra-Wasserburg Plot:')
        worksheet.write(0, 26, '238U/206Pb')
        worksheet.write(0, 27, '2s error')
        worksheet.write(0, 28, '207Pb/206Pb')
        worksheet.write(0, 29, '2s error')
        worksheet.write(0, 30, 'Plotting purposes:')
        worksheet.write(0, 31, '207Pb/235U')
        worksheet.write(0, 32, '2s error')
        worksheet.write(0, 33, '206Pb/238U')
        worksheet.write(0, 34, '2s error')
        worksheet.write(0, 35, 'rho')
        worksheet.write(0, 36, '208Pb/232Th')
        worksheet.write(0, 37, '2s error')
        worksheet.write(0, 38, '208Pb/206Pb')
        worksheet.write(0, 39, '2s error')
        
        
        for row in range(len(result_cal_age)):
            for col in range(40):
                worksheet.write(row+1,col,result_cal_age[row][col])
            
            
        
        now = time.strftime("%H%M%S",time.localtime(time.time()))     
        workbook.save('cal_age_result_L'+now+'.xls')
        
        logging.info('The calculation was successful.The file saved in:%s',outputpath)
       
    except ValueError:
        
        tk.messagebox.showinfo(title='Information！', message='result_all.csv"file not found or  the format is incorrect！')
        logging.critical('“result_all.csv" were not found or the format is incorrect.')  
    except NameError:
        
        tk.messagebox.showinfo(title='Information！', message='Output directory or standard setting error！')
        logging.error('The output directory is not set.') 
    except OSError:
        tk.messagebox.showinfo(title='Information！', message='"result_all.csv" not found！')
        logging.error('"result_all.csv" not found')
        print('"result_all.csv" not found')
        
    tk.messagebox.showinfo(title='Congratulations！', message='The age calculation was successful！')
      
         

def dataprocess():
    #print(sampleslist)
    logging.debug('Correction for all samples.' )

       
    result_outstd=[]
    result_outsamples=[]
    i=1
    scale=len(sampleslist)
    wrong_samples=[]
    def check():
        fa=[]
        numn=np.load('Time_setting.npy',allow_pickle='True').item() 
        for name in numn.keys():
            if numn[name][3]-numn[name][2]<0 or numn[name][4]-numn[name][2]<0:
                fa.append(name)
            elif numn[name][3]+1>numn[name][4]-1:
                numn[name][3]=numn[name][4]
                numn[name][4]=numn[name][3]
            else:
                pass        
        np.save('Time_setting.npy',numn)
        return fa
    wrong_samples=check()
        
    
    try:
        numn=np.load('Time_setting.npy',allow_pickle='True').item() 
        for name in sampleslist.keys():
            
            b0=numn[name][1]
            b1=numn[name][2]
            s0=numn[name][3]
            s1=numn[name][4]
            start_b=int(b0/Timeinternal)
            end_b=int(b1/Timeinternal)
            start_s=int(s0/Timeinternal)
            end_s=int(s1/Timeinternal)     
           
            
            if stdcor.get()==0:          
        
                if sampleslist[name] in Standard_names.keys():
                    
                    Npname_Time_b=name+'_x.npy'
                    Npname_Hg202_b=name+'_y1.npy'
                    Npname_Hg204_b=name+'_y2.npy'
                    Npname_Pb206_b=name+'_y3.npy'
                    Npname_Pb207_b=name+'_y4.npy'
                    Npname_Pb208_b=name+'_y5.npy'
                    Npname_Th232_b=name+'_y6.npy'
                    Npname_U238_b=name+'_y7.npy'
                    
                    os.chdir(outputpath)
                    Time_b=np.load(Npname_Time_b).T[start_b:end_b]
                    Hg202_b=np.load(Npname_Hg202_b).T[start_b:end_b]
                    Hg204_b=np.load(Npname_Hg204_b).T[start_b:end_b]
                    Pb206_b=np.load(Npname_Pb206_b).T[start_b:end_b]
                    Pb207_b=np.load(Npname_Pb207_b).T[start_b:end_b]
                    Pb208_b=np.load(Npname_Pb208_b).T[start_b:end_b]
                    Th232_b=np.load(Npname_Th232_b).T[start_b:end_b]
                    U238_b=np.load(Npname_U238_b).T[start_b:end_b]
                    
                    Time_s=np.load(Npname_Time_b).T[start_s:end_s]
                    Hg202_s=np.load(Npname_Hg202_b).T[start_s:end_s]
                    Hg204_s=np.load(Npname_Hg204_b).T[start_s:end_s]
                    Pb206_s=np.load(Npname_Pb206_b).T[start_s:end_s]
                    Pb207_s=np.load(Npname_Pb207_b).T[start_s:end_s]
                    Pb208_s=np.load(Npname_Pb208_b).T[start_s:end_s]
                    Th232_s=np.load(Npname_Th232_b).T[start_s:end_s]
                    U238_s=np.load(Npname_U238_b).T[start_s:end_s]         
                    
                    
                    result_std=Std_method(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,\
                                          Hg202_s,Hg204_s,Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s)
                    result_outstd.append(result_std)
                
                
                else:                   
                    
                    Npname_Time_b=name+'_x.npy'
                    Npname_Hg202_b=name+'_y1.npy'
                    Npname_Hg204_b=name+'_y2.npy'
                    Npname_Pb206_b=name+'_y3.npy'
                    Npname_Pb207_b=name+'_y4.npy'
                    Npname_Pb208_b=name+'_y5.npy'
                    Npname_Th232_b=name+'_y6.npy'
                    Npname_U238_b=name+'_y7.npy'
                    
                    os.chdir(outputpath)
                    Time_b=np.load(Npname_Time_b).T[start_b:end_b]
                    Hg202_b=np.load(Npname_Hg202_b).T[start_b:end_b]
                    Hg204_b=np.load(Npname_Hg204_b).T[start_b:end_b]
                    Pb206_b=np.load(Npname_Pb206_b).T[start_b:end_b]
                    Pb207_b=np.load(Npname_Pb207_b).T[start_b:end_b]
                    Pb208_b=np.load(Npname_Pb208_b).T[start_b:end_b]
                    Th232_b=np.load(Npname_Th232_b).T[start_b:end_b]
                    U238_b=np.load(Npname_U238_b).T[start_b:end_b]
                    
                    Time_s=np.load(Npname_Time_b).T[start_s:end_s]
                    Hg202_s=np.load(Npname_Hg202_b).T[start_s:end_s]
                    Hg204_s=np.load(Npname_Hg204_b).T[start_s:end_s]
                    Pb206_s=np.load(Npname_Pb206_b).T[start_s:end_s]
                    Pb207_s=np.load(Npname_Pb207_b).T[start_s:end_s]
                    Pb208_s=np.load(Npname_Pb208_b).T[start_s:end_s]
                    Th232_s=np.load(Npname_Th232_b).T[start_s:end_s]
                    U238_s=np.load(Npname_U238_b).T[start_s:end_s]
                      
                    
                    
                       
                    result_samples=samples_method(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,\
                                                  Pb208_b,Th232_b,U238_b,Hg202_s,Hg204_s,Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s)
                    result_outsamples.append(result_samples)
            else:
                Npname_Time_b=name+'_x.npy'
                Npname_Hg202_b=name+'_y1.npy'
                Npname_Hg204_b=name+'_y2.npy'
                Npname_Pb206_b=name+'_y3.npy'
                Npname_Pb207_b=name+'_y4.npy'
                Npname_Pb208_b=name+'_y5.npy'
                Npname_Th232_b=name+'_y6.npy'
                Npname_U238_b=name+'_y7.npy'
                
                os.chdir(outputpath)
                Time_b=np.load(Npname_Time_b).T[start_b:end_b]
                Hg202_b=np.load(Npname_Hg202_b).T[start_b:end_b]
                Hg204_b=np.load(Npname_Hg204_b).T[start_b:end_b]
                Pb206_b=np.load(Npname_Pb206_b).T[start_b:end_b]
                Pb207_b=np.load(Npname_Pb207_b).T[start_b:end_b]
                Pb208_b=np.load(Npname_Pb208_b).T[start_b:end_b]
                Th232_b=np.load(Npname_Th232_b).T[start_b:end_b]
                U238_b=np.load(Npname_U238_b).T[start_b:end_b]
                
                Time_s=np.load(Npname_Time_b).T[start_s:end_s]
                Hg202_s=np.load(Npname_Hg202_b).T[start_s:end_s]
                Hg204_s=np.load(Npname_Hg204_b).T[start_s:end_s]
                Pb206_s=np.load(Npname_Pb206_b).T[start_s:end_s]
                Pb207_s=np.load(Npname_Pb207_b).T[start_s:end_s]
                Pb208_s=np.load(Npname_Pb208_b).T[start_s:end_s]
                Th232_s=np.load(Npname_Th232_b).T[start_s:end_s]
                U238_s=np.load(Npname_U238_b).T[start_s:end_s]
                  
                
                
                   
                result_samples=samples_method(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,\
                                              Hg202_s,Hg204_s,Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s)
                result_outsamples.append(result_samples)
            


        result_outstd.extend(result_outsamples)
        result_outstd.sort(key=lambda x: x[0],reverse=False)

        with open("result_all.csv", "w", newline='') as s:
            writer=csv.writer(s)
            writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s',\
                             '208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s','Trace element ','U(cps)','Th(cps)',\
                             'Pb208(cps)','Pb207(cps)','Pb206(cps)'])
            for r in result_outstd:
                writer.writerow(r)
        

  
    
    
        '''fileSubfix =['npy','png']
    for parent,dirnames,filenames in os.walk(outputpath):            
        for filename in filenames:
            if filename.split('.')[-1] in fileSubfix:
                os.remove(filename)'''
            
        tk.messagebox.showinfo(title='Congratulations！', message='Successful！')
    

    except NameError:
        logging.error('Standard or dwell times is not set !')
       
        tk.messagebox.showinfo(title='Error！', message='Standard or dwell times is not set !')

    except KeyError:
        print('Wrong！Dwell times are too short！')
        logging.critical('warning:Dwell times are too short.')
        tk.messagebox.showinfo(title='Error！', message='Dwell times are too short！')
    except ValueError:        
        print('Incorrect setting of integration！')
        print(wrong_samples)       
        
        logging.critical('warning:Incorrect setting of integration:')
        if wrong_samples:
            tk.messagebox.showinfo(title='ValueError！', message='Incorrect setting of integration！'+wrong_samples[0])
        else:
            tk.messagebox.showinfo(title='ValueError！', message='Incorrect setting of integration！')
            
        
    except Exception as e:
        logging.error('Unknown erro:%s',str(e))
        print('Unknown erro！')
        tk.messagebox.showinfo(title='Error！', message='Unknown erro！')
        print(str(e))
        
   

    


def samples_method(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,Hg202_s,Hg204_s,\
                   Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s):

    N383=0.0
    num=np.load('Time_setting.npy',allow_pickle='True').item()
    b0=num[name][1]
    b1=num[name][2]
    s0=num[name][3]
    s1=num[name][4]
    start_b=int(b0/Timeinternal)
    end_b=int(b1/Timeinternal)
    start_s=int(s0/Timeinternal)
    end_s=int(s1/Timeinternal)

    
    #  Calculation for samples
    Hg204_Hg202_s=Hg204_s/Hg202_s
    Hg204cal_s=(np.average(Hg202_s)-np.average(Hg202_s))*0.229883    
    m204_s=np.where(Hg204cal_s<(np.std(Hg204_s)/sqrt(end_s-start_s)),0,Hg204cal_s)    
    Pb204_s=np.average(Hg204_s)-np.average(Hg204_b)
    
    U238_T=(np.average(U238_s)-np.average(U238_b))
    Th232_T=(np.average(Th232_s)-np.average(Th232_b))
    Pb208_T=(np.average(Pb208_s)-np.average(Pb208_b))
    Pb207_T=(np.average(Pb207_s)-np.average(Pb207_b))
    Pb206_T=(np.average(Pb206_s)-np.average(Pb206_b))
    
    Pb207_Pb206_s=(Pb207_s-np.average(Pb207_b)-N383*R8)/(Pb206_s-np.average(Pb206_b)-N383*Q8)    
    Pb206_U238_s=(Pb206_s-np.average(Pb206_b)-N383*Q8)/(U238_s-np.average(U238_b))        
    Pb208_Th232_s=(Pb208_s-np.average(Pb208_b)-N383*Q8)/(Th232_s-np.average(Th232_b))    
    Pb208_Pb206_s=(Pb208_s-np.average(Pb208_b)-N383*Q8)/(Pb206_s-np.average(Pb206_b))    
    Th232_Pb206_s=Th232_s/Pb206_s    
    try:
        Pb208_Pb204_s=(Pb208_s-np.average(Pb208_b))/Pb204_s
    except:
        pass
    
    
    #filtering of outliers
    
    

    
    Pb207_Pb206_f=np.where(np.abs(Pb207_Pb206_s-np.average(Pb207_Pb206_s))<2*np.std(Pb207_Pb206_s),Pb207_Pb206_s,nan)
    
    
    Pb206_U238_f=np.where(np.abs(Pb206_U238_s-np.average(Pb206_U238_s))<2*np.std(Pb206_U238_s),Pb206_U238_s,nan)

    
    
    
       
    Pb208_Th232_f=np.where(np.abs(Pb208_Th232_s-np.average(Pb208_Th232_s))<2*np.std(Pb208_Th232_s),Pb208_Th232_s,nan)
    
    Pb208_Pb206_f=np.where(np.abs(Pb208_Pb206_s-np.average(Pb208_Pb206_s))<2*np.std(Pb208_Pb206_s),Pb208_Pb206_s,nan)
    
    
    
    Th232_Pb206_f=np.where(np.abs(Th232_Pb206_s-np.average(Th232_Pb206_s))<2*np.std(Th232_Pb206_s),Th232_Pb206_s,nan)
    
    Pb208_Pb204_f=np.where(np.abs(Pb208_Pb204_s-np.average(Pb208_Pb204_s))<2*np.std(Pb208_Pb204_s),Pb208_Pb204_s,nan)
    
    
    Pb207_U235_c=np.nanmean(Pb207_Pb206_f)*np.nanmean(Pb206_U238_f)*137.88
    
    if ele.get()==0:
        
        No=int(name.split('_')[-1].split('.')[0])
    else:
        
        No=name
                        
    result_samples=[No,name,sampleslist[name],np.nanmean(Pb207_Pb206_f),\
                    2*np.nanstd(Pb207_Pb206_f)/sqrt(len(Pb207_Pb206_f)-np.count_nonzero(np.isnan(Pb207_Pb206_f))),
                    np.nanmean(Pb206_U238_f),\
                    2*np.nanstd(Pb206_U238_f)/sqrt(len(Pb206_U238_f)-np.count_nonzero(np.isnan(Pb206_U238_f))),
                    Pb207_U235_c,\
                    sqrt(pow(Pb207_U235_c*(((2*np.nanstd(Pb206_U238_f)/sqrt(len(Pb206_U238_f)-np.count_nonzero(np.isnan(Pb206_U238_f)))))/np.nanmean(Pb206_U238_f))*100/100,2)+pow(Pb207_U235_c*(((2*np.nanstd(Pb207_Pb206_f)/sqrt(len(Pb207_Pb206_f)-np.count_nonzero(np.isnan(Pb207_Pb206_f)))))/np.nanmean(Pb207_Pb206_f))*100/100,2)),
                    np.nanmean(Pb208_Th232_f),\
                    2*np.nanstd(Pb208_Th232_f)/sqrt(len(Pb208_Th232_f)-np.count_nonzero(np.isnan(Pb208_Th232_f))),
                    np.nanmean(Pb208_Pb206_f),\
                    2*np.nanstd(Pb208_Pb206_f)/sqrt(len(Pb208_Pb206_f)-np.count_nonzero(np.isnan(Pb208_Pb206_f))),
                    np.nanmean(Th232_Pb206_f),\
                    2*np.nanstd(Th232_Pb206_f)/sqrt(len(Th232_Pb206_f)-np.count_nonzero(np.isnan(Th232_Pb206_f))),
                    np.nanmean(Pb208_Pb204_f),\
                    2*np.nanstd(Pb208_Pb204_f)/sqrt(len(Pb208_Pb204_f)-np.count_nonzero(np.isnan(Pb208_Pb204_f)))
                    ,'------',U238_T,Th232_T,Pb208_T,Pb207_T,Pb206_T]
    
    return result_samples
   
    
    
def Std_method(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,Hg202_s,Hg204_s,Pb206_s,Pb207_s,\
               Pb208_s,Th232_s,U238_s):
    
    num=np.load('Time_setting.npy',allow_pickle='True').item()
    b0=num[name][1]
    b1=num[name][2]
    s0=num[name][3]
    s1=num[name][4]
    start_b=int(b0/Timeinternal)
    end_b=int(b1/Timeinternal)
    start_s=int(s0/Timeinternal)
    end_s=int(s1/Timeinternal)

    def Std_process(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,Hg202_s,Hg204_s,\
                    Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s,N383):
        num=np.load('Time_setting.npy',allow_pickle='True').item()
        b0=num[name][1]
        b1=num[name][2]
        s0=num[name][3]
        s1=num[name][4]
        start_b=int(b0/Timeinternal)
        end_b=int(b1/Timeinternal)
        start_s=int(s0/Timeinternal)
        end_s=int(s1/Timeinternal)
        # Calculation for samples
        Hg204_Hg202_s=Hg204_s/Hg202_s
        Hg204cal_s=(np.average(Hg202_s)-np.average(Hg202_s))*0.229883
        
        m204_s=np.where(Hg204cal_s<(np.std(Hg204_s)/sqrt(end_s-start_s)),0,Hg204cal_s)
        
        Pb204_s=np.average(Hg204_s)-np.average(Hg204_b)
        
        U238_T=np.average(U238_s)-np.average(U238_b)
        Th232_T=np.average(Th232_s)-np.average(Th232_b)
        Pb208_T=(np.average(Pb208_s)-np.average(Pb208_b))
        Pb207_T=(np.average(Pb207_s)-np.average(Pb207_b))
        Pb206_T=(np.average(Pb206_s)-np.average(Pb206_b))
        
        Pb207_Pb206_s=(Pb207_s-np.average(Pb207_b)-N383*R8)/(Pb206_s-np.average(Pb206_b)-N383*Q8)
        
        Pb206_U238_s=(Pb206_s-np.average(Pb206_b)-N383*Q8)/(U238_s-np.average(U238_b))
            
        Pb208_Th232_s=(Pb208_s-np.average(Pb208_b)-N383*Q8)/(Th232_s-np.average(Th232_b))
        
        Pb208_Pb206_s=(Pb208_s-np.average(Pb208_b)-N383*Q8)/(Pb206_s-np.average(Pb206_b))
        
        Th232_Pb206_s=Th232_s/Pb206_s
        
        Pb208_Pb204_s=(Pb208_s-np.average(Pb208_b))/Pb204_s
        
        
        #filtering of outliers
        
        
        x=np.arange(len(Pb206_s))
        Pb207_Pb206_f=np.where(np.abs(Pb207_Pb206_s-np.average(Pb207_Pb206_s))<2*np.std(Pb207_Pb206_s),Pb207_Pb206_s,nan)
        
        Pb206_U238_f=np.where(np.abs(Pb206_U238_s-np.average(Pb206_U238_s))<2*np.std(Pb206_U238_s),Pb206_U238_s,nan)
         
        
        
           
        Pb208_Th232_f=np.where(np.abs(Pb208_Th232_s-np.average(Pb208_Th232_s))<2*np.std(Pb208_Th232_s),Pb208_Th232_s,nan)
        
        Pb208_Pb206_f=np.where(np.abs(Pb208_Pb206_s-np.average(Pb208_Pb206_s))<2*np.std(Pb208_Pb206_s),Pb208_Pb206_s,nan)
        
        
        
        Th232_Pb206_f=np.where(np.abs(Th232_Pb206_s-np.average(Th232_Pb206_s))<2*np.std(Th232_Pb206_s),Th232_Pb206_s,nan)
        
        Pb208_Pb204_f=np.where(np.abs(Pb208_Pb204_s-np.average(Pb208_Pb204_s))<2*np.std(Pb208_Pb204_s),Pb208_Pb204_s,nan)
        
        Pb207_U235_c=np.nanmean(Pb207_Pb206_f)*np.nanmean(Pb206_U238_f)*137.88
        

        corr_Pb207_206=np.nanmean(Pb207_Pb206_f)
        
        if ele.get()==0:
            No=int(name.split('_')[-1].split('.')[0])
        else:
            No=name
        result_std=[No,name,sampleslist[name],np.nanmean(Pb207_Pb206_f),\
                    2*np.nanstd(Pb207_U235_c/Pb206_U238_f/137.88)/sqrt(len(Pb207_U235_c/Pb206_U238_f/137.88)-np.count_nonzero(np.isnan(Pb207_U235_c/Pb206_U238_f/137.88))),
                    np.nanmean(Pb206_U238_f),\
                    2*np.nanstd(Pb206_U238_f)/sqrt(len(Pb206_U238_f)-np.count_nonzero(np.isnan(Pb206_U238_f))),
                    Pb207_U235_c,\
                    sqrt(pow(Pb207_U235_c*(((2*np.nanstd(Pb206_U238_f)/sqrt(len(Pb206_U238_f)-np.count_nonzero(np.isnan(Pb206_U238_f)))))/np.nanmean(Pb206_U238_f))*100/100,2)+pow(Pb207_U235_c*(((2*np.nanstd(Pb207_Pb206_f)/sqrt(len(Pb207_Pb206_f)-np.count_nonzero(np.isnan(Pb207_Pb206_f)))))/np.nanmean(Pb207_Pb206_f))*100/100,2)),
                    np.nanmean(Pb208_Th232_f),\
                    2*np.nanstd(Pb208_Th232_f)/sqrt(len(Pb208_Th232_f)-np.count_nonzero(np.isnan(Pb208_Th232_f))),
                    np.nanmean(Pb208_Pb206_f),\
                    2*np.nanstd(Pb208_Pb206_f)/sqrt(len(Pb208_Pb206_f)-np.count_nonzero(np.isnan(Pb208_Pb206_f))),
                    np.nanmean(Th232_Pb206_f),\
                    2*np.nanstd(Th232_Pb206_f)/sqrt(len(Th232_Pb206_f)-np.count_nonzero(np.isnan(Th232_Pb206_f))),
                    np.nanmean(Pb208_Pb204_f),\
                    2*np.nanstd(Pb208_Pb204_f)/sqrt(len(Pb208_Pb204_f)-np.count_nonzero(np.isnan(Pb208_Pb204_f)))
                    ,'------',U238_T,Th232_T,Pb208_T,Pb207_T,Pb206_T]
        
        return result_std
    
    Pb207_Pb206_s=np.ones(end_s-start_s)*P382
    N383=(Pb207_Pb206_s*Pb206_s-np.average(Pb206_b)*Pb207_Pb206_s-Pb207_s+np.average(Pb207_b))/(Q8*Pb207_Pb206_s-R8)
    
         
    result_std=Std_process(name,Hg202_b,Hg204_b,Pb206_b,Pb207_b,Pb208_b,Th232_b,U238_b,Hg202_s,Hg204_s,\
                           Pb206_s,Pb207_s,Pb208_s,Th232_s,U238_s,N383)


            

        
         
    return  result_std

def showplt(event):
    
    index = theLB.curselection()
    var=theLB.get(index)
    #print(var)
    x,y1,y2,y3,y4,y5,y6,y7=loaddata(var,isoname)
    numn=np.load('Time_setting.npy',allow_pickle='True').item()
    #print(numn)
    
    
    
       
    
    
    
    plt.plot(x, y1,label='202Hg')
    plt.plot(x, y2,label='204Pb')
    plt.plot(x, y3,label='206Pb',ls='--')
    plt.plot(x, y4,label='207Pb')
    plt.plot(x, y5,label='208Pb')
    plt.plot(x, y6,label='232Th')
    plt.plot(x, y7,label='238U',ls='-.')
    ax=plt.gca()
    miloc=plt.MultipleLocator(1)
    ax.xaxis.set_minor_locator(miloc)
    
    ax.set_yscale('log')
    plt.legend()
    plt.grid(axis='x')
    
    

    
    
    b0=numn[var][1]
    b1=numn[var][2]
    s0=numn[var][3]
    s1=numn[var][4]




    
    plt.axvline(b0,color='b',linestyle='--',label='bcg_s',linewidth=3)
    plt.axvline(b1,color='b',linestyle='--',label='bcg_e',linewidth=3)
    plt.axvline(s0,color='black',linestyle='--',label='sig_s',linewidth=3)
    plt.axvline(s1,color='black',linestyle='--',label='sig_e',linewidth=3)
    
    plt.title(var.split('.')[0]+'  :{'+sampleslist[var]+'}')
    plt.tight_layout()
    fname=var.split('.')[0]        
    save=plt.savefig(outputpath+'//'+fname+'.png')
    plt.clf()
    plt.xlabel("Time(s)")
    plt.ylabel("Counts (cps)")
    img_open=Image.open(outputpath+'//'+fname+'.png')
    img_open2=img_open.resize((400,300))
    img=ImageTk.PhotoImage(img_open2)
    l=Label(frmRB)
    l.grid(row =5, column =5,padx=5,pady=10)
    l.config(image=img)
    l.image=img
    
    
   
 

 

    
def instructure0():
    global sampleslist
    global isoname
    sampleslist={}
    os.chdir(outputpath)

    
   
    '''with open("samples.csv", "a", newline='') as s:
        
        writer=csv.writer(s)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])'''
    with open("result_all.csv", "a", newline='') as s:
        
        writer=csv.writer(s)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])      

    '''with open("result_all.csv", "a", newline='') as g:
        writer=csv.writer(g)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s',\
                         '208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])'''      


    samplename=[]
    
    global Numbers
    global Timeinternal
    global fig_name
    
    num={}
    window = tk.Tk()
    window.title('Loading......')
    window.geometry('630x150')
    tk.Label(window, text='Progress of loading:', ).place(x=50, y=60)
    
    canvas = tk.Canvas(window, width=465, height=22, bg="white")
    loading_state= StringVar()
    w=Canvas(window, width=200, height=22, bg="pink")
    canvas.place(x=110, y=60)
    w.place(x=250, y=90)
       
     
   
        
    
    for root,dirs,files in os.walk(inputpath):
        i=1
        scale=len(files)
        
        print('Start Loading....'.center(scale//2,'='))
        fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")
        
        n = 465 / scale 
        files_list=[]

        for name in files:
            if os.path.splitext(name)[1]=='.csv':
                files_list.append(name)
        for name in files_list:
            w.create_text(90,10,text=f'Loading.... {name} ')
            a='*'*i
            b='.'*(scale-i)
            c=(i/scale)*100
            isoname=['Time','202Hg','204Pb','206Pb','207Pb','208Pb','232Th','238U']
            x,y1,y2,y3,y4,y5,y6,y7=loaddata(name,isoname)
            Numbers=len(x)           
                           
                
            Timeinternal=(x[Numbers-1]-x[0])/Numbers
            
            b0=bcg_from
            b1=bcg_to
            y8=[]

            bcg_s=int(b0/Timeinternal)
            bcg_e=int(b1/Timeinternal)
            
            for i in range(len(x)):
                sum_a=y1[i]+y2[i]+y3[i]+y4[i]+y5[i]+y6[i]+y7[i]
                y8.append(sum_a)
            signal_bcg=np.array(y8)[bcg_s:bcg_e]
                
                
            signal_bcg=np.where(np.abs(signal_bcg-np.average(signal_bcg))<2*np.std(signal_bcg),signal_bcg,np.average(signal_bcg))   

            bcg=signal_bcg.mean()
        
      
            std=np.std(y8[bcg_s:bcg_e])
            sdmul=multi
            ind = [True if value > bcg+sdmul *std else False for value in y8]
            ind2 = ind[1:]
            ind2.append(False)


            #print(ind)
            #print(ind2)
            index = [i for i in range(0, len(ind)) if ind[i] != ind2[i]]

                    
            
            starts = [index[i] for i in range(len(index)) if i % 2 == 0]
            ends = [index[i] for i in range(len(index)) if i % 2 != 0]
            try:
                #print(name,':','s',starts,'e',ends)
                if len(ends)>1 and len(starts)>1:
                    for i in range(len(starts)):
                        posi=0
                        if starts[i]<((b1)/Timeinternal):
                            s0=float(starts[posi+1])*abs(Timeinternal)  
                        else:
                            s0=float(starts[posi])*abs(Timeinternal)
                        pos=posi+1
                    for j in range(len(ends)):
                        posj=0
                        if ends[j]<((b1)/Timeinternal):
                            s1=float(ends[posj+1])*abs(Timeinternal)
                            
                        else:
                            s1=float(ends[posj])*abs(Timeinternal)
                        posj=posj+1
                    

                else:
                    s1=float(ends[0])*abs(Timeinternal)
                    s0=float(starts[0])*abs(Timeinternal)              
            except:
                s0=30
                s1=60
            Diff=(s1-1)-(s0+1)
            if Diff>0:
                num[name]=[Numbers,b0,b1,s0+1,s1-1]
                
            else:num[name]=[Numbers,b0,b1,s1+1,s0-1]      
    

            np.save(outputpath+'//'+name+'_x',x)
            np.save(outputpath+'//'+name+'_y1',y1)
            np.save(outputpath+'//'+name+'_y2',y2)
            np.save(outputpath+'//'+name+'_y3',y3)
            np.save(outputpath+'//'+name+'_y4',y4)
            np.save(outputpath+'//'+name+'_y5',y5)
            np.save(outputpath+'//'+name+'_y6',y6)
            np.save(outputpath+'//'+name+'_y7',y7)
            
            info=open(inputpath+'\\'+name)
            headreader=csv.reader(info)
            for j,row in enumerate(headreader):
                if j==0:
                    ss=str(row[0])
                    ss=ss.split(':')
                    #If standard  name  contain '-'  
                    #ss=re.split(':|-',ss)
                    #ss=ss[0].split('-')
                    sampleslist[name]=ss[0]
            i=i+1      
   
 

        
            
            
            n = n + 465 / scale
           
            
            
            canvas.coords(fill_line, (0, 0, n, 60))
            
            
            window.update()
            time.sleep(0) 
            w.delete("all")
            #print("\r{:^3.0f}%[{}->{}]".format(c,a,b),end='')
    np.save(outputpath+'//'+'Time_setting',num)
    window.destroy()
    '''if max(num.values())-min(num.values())>=5:
        print()
        logging.critical('Warning notice: The original data file is abnormal!Plese check %s',str(min(num,key=num.get)))
        tk.messagebox.showinfo(title='Warning ！', message='he original data file is abnormal!Plese check ：'+str(min(num,key=num.get)))
        print('he original data file is abnormal!Plese check ：',min(num,key=num.get))
  
    else:
    '''
    logging.info("Loading compeleted.")   
    print()
    print('Loading compeleted'.center(scale//2,'='))
    print()

    tk.messagebox.showinfo(title='Information！', message='Loading completed!Total '+str(len(files_list))+'files have been loaded.')
    logging.info("Total %s files have been loaded.",str(len(files_list)))  
        
   

    global theLB,sc
    theLB=Listbox(frmLB,selectmode=SINGLE,height=11)
    theLB.grid(row =6,column =0,rowspan=4,columnspan=3,pady=2,padx=5)
    sc = tkinter.Scrollbar(frmLB,command=theLB.yview)
    sc.grid(row =6,column =3,rowspan=4,pady=2,sticky='ns') 
   
    theLB.bind("<<ListboxSelect>>", showplt)
    theLB.configure(yscrollcommand = sc.set)
    
    
    
    for item in sampleslist.keys():
        fig_name=item
        theLB.insert(END,fig_name)
    def selectall():
        theLB.select_set(0,END)
        
    def cancleall():
        theLB.selection_clear(0,END)
    def stdbutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] ==standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                print('Key Error')
                logging.error('Standard is not set!')

    def sambutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] !=standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                
                print()
            except NameError:
                print('Standard is not set！')
                logging.error('Standard is not set!')
            except Exception as e:
                print(str(e))
                logging.error('Unknown erro:%s',str(e))
        
   
    theButton = Button(frmLB,text='Delete',\
                   command=lambda x=theLB:x.delete(ACTIVE))
  
    theButton.grid(row =11,column =2,pady=5)
    showButton=Button(frmLB,text='All',command=selectall)
    showButton.grid(row =11,column =0,pady=5)
 
    cancleButton=Button(frmLB,text='Cancle',command=cancleall)
    cancleButton.grid(row =11,column =1,pady=5)

   
    
    #window.mainloop()
    return fig_name

        
    
def instructure1():
    global sampleslist
    global isoname
    sampleslist={}
    os.chdir(outputpath)
    
   
    with open("samples.csv", "a", newline='') as s:
        
        writer=csv.writer(s)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])
            
    with open("std.csv", "a", newline='') as f:
        writer=csv.writer(f)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])
    with open("result_all.csv", "a", newline='') as g:
        writer=csv.writer(g)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])      


    samplename=[]
    
    global Numbers
    global Timeinternal
    global fig_name
    
    num={}
    root = tk.Tk()
    root.withdraw()
    default_dir = inputpath
    file_path=tk.filedialog.askopenfilename(title=u'Choose the LIST file', initialdir=(os.path.expanduser(default_dir)))
    data = xlrd.open_workbook(file_path)
    data.sheet_names()
    try:
        table = data.sheet_by_name('Sheet1')
    except Exception as e:
        print(str(e))
        tk.messagebox.showinfo(title='Information！', message='Please check LIST file！')
        logging.error("Error file：%s ",str(e))
    Sampleslist1={}

   
    

    
    window = tk.Tk()
    window.title('Loading.......')
    window.geometry('630x150')
    tk.Label(window, text='Progress of loading:', ).place(x=50, y=60)
    canvas = tk.Canvas(window, width=465, height=22, bg="white")
    canvas.place(x=110, y=60)   
    w=Canvas(window, width=200, height=22, bg="pink")   
    w.place(x=250, y=90)
       
    for i in range(table.nrows):
        if isinstance(table.cell(i,1).value ,float):
            Sampleslist1[table.cell(i,0).value]='%d'%(table.cell(i,1).value )
        else:
             Sampleslist1[table.cell(i,0).value]='%s'%(table.cell(i,1).value )
            
    i=1
    files_list=[]

    for root,dirs,files in os.walk(inputpath):
        
        for name in files:            
            if os.path.splitext(name)[1]=='.csv':
                files_list.append(name)
    scale=len(files_list)
    print('Start Loading....'.center(scale//2,'='))
    fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")
        
    n = 465 / scale  # 465是矩形填充满的次数    
      
        
    for name in files_list:
        w.create_text(90,10,text=f'Loading.... {name} ')  
        a='*'*i
        b='.'*(scale-i)
        c=(i/scale)*100
        #print(name)
        isoname=['时间 [s]','202','204','206','207','208','232','238']
        x,y1,y2,y3,y4,y5,y6,y7=loaddata(name,isoname)
         #Save the signal and background time for automatic identification

        Numbers=len(x)
        Timeinternal=(x[Numbers-1]-x[0])/Numbers

        
        b0=bcg_from
        b1=bcg_to
        y8=[]

        bcg_s=int(b0/Timeinternal)
        bcg_e=int(b1/Timeinternal)
        
        for i in range(len(x)):
            sum_a=y1[i]+y2[i]+y3[i]+y4[i]+y5[i]+y6[i]+y7[i]
            y8.append(sum_a)
        signal_bcg=np.array(y8)[bcg_s:bcg_e]
            
            
        signal_bcg=np.where(np.abs(signal_bcg-np.average(signal_bcg))<2*np.std(signal_bcg),signal_bcg,np.average(signal_bcg))   

        bcg=signal_bcg.mean()
    
  
        std=np.std(y8[bcg_s:bcg_e])
        sdmul=multi
        ind = [True if value > bcg+sdmul *std else False for value in y8]
        ind2 = ind[1:]
        ind2.append(False)


        #print(ind)
        #print(ind2)
        index = [i for i in range(0, len(ind)) if ind[i] != ind2[i]]

                
        s0=30
        s1=60
        starts = [index[i] for i in range(len(index)) if i % 2 == 0]
        ends = [index[i] for i in range(len(index)) if i % 2 != 0]
        #print(name,':','s',starts,'e',ends)
        try :
            if len(ends)>1 and len(starts)>1:
                for i in range(len(starts)):
                    posi=0
                    if starts[i]<((b1)/Timeinternal):
                        s0=float(starts[posi+1])*abs(Timeinternal)  
                    else:
                        s0=float(starts[posi])*abs(Timeinternal)
                    pos=posi+1
                for j in range(len(ends)):
                    posj=0
                    if ends[j]<((b1)/Timeinternal):
                        s1=float(ends[posj+1])*abs(Timeinternal)
                        
                    else:
                        s1=float(ends[posj])*abs(Timeinternal)
                    posj=posj+1
                
              
            else:
                s1=float(ends[0])*abs(Timeinternal)
                s0=float(starts[0])*abs(Timeinternal)              
        except Exception as e:
            print(str(e))
            
        Diff=(s1-1)-(s0+1)
        if Diff>0:
            num[name]=[Numbers,b0,b1,s0+1,s1-1]
            
        else:num[name]=[Numbers,b0,b1,s1+1,s0-1]      
    
        


        np.save(outputpath+'//'+name+'_x',x)
        np.save(outputpath+'//'+name+'_y1',y1)
        np.save(outputpath+'//'+name+'_y2',y2)
        np.save(outputpath+'//'+name+'_y3',y3)
        np.save(outputpath+'//'+name+'_y4',y4)
        np.save(outputpath+'//'+name+'_y5',y5)
        np.save(outputpath+'//'+name+'_y6',y6)
        np.save(outputpath+'//'+name+'_y7',y7)
        #info=open(inputpath+'\\'+name.replace('.csv','.d')+name)
        sampleslist[name]=Sampleslist1[name.split('.')[0]]
        i=i+1
        n = n + 465 / scale
           
        canvas.coords(fill_line, (0, 0, n, 60))
        window.update()
        time.sleep(0)  
        #print("\r{:^3.0f}%[{}->{}]".format(c,a,b),end='')
        w.delete("all")  
    np.save(outputpath+'//'+'Time_setting',num)
    window.destroy()
    tk.messagebox.showinfo(title='Information！', message='Loading compeleted！Total '+str(len(sampleslist))+'files')
    '''if max(num.values()[0])-min(num.values()[0])>=5:
        print()
        logging.critical('Warning notice: The original data file is abnormal!Plese check %s',str(min(num,key=num.get)))

        tk.messagebox.showinfo(title='Warning！', message='The original data file is abnormal!Plese check'+str(min(num,key=num.get)))
        print('The original data file is abnormal!Plese check：',min(num,key=num.get))
        
    else:
        logging.info("Loading compeleted.")   
        print()
        print('Loading compeleted'.center(scale//2,'='))
        print()
        logging.info("Total %s files have been loaded.",str(len(files_list)))  
        tk.messagebox.showinfo(title='Information！', message='Loading compeleted！Total '+str(len(sampleslist))+'files')
      
      ''' 

   
    global theLB,sc
    theLB=Listbox(frmLB,selectmode=SINGLE,height=11)
    theLB.grid(row =6,column =0,rowspan=4,columnspan=3,pady=2,padx=5)
    sc = tkinter.Scrollbar(frmLB,command=theLB.yview)
    sc.grid(row =6,column =3,rowspan=4,pady=2,sticky='ns') 
    theLB.bind("<<ListboxSelect>>", showplt)
    theLB.configure(yscrollcommand = sc.set)
    for item in sampleslist.keys():
        fig_name=item
        theLB.insert(END,fig_name)
    def selectall():
        theLB.select_set(0,END)
        
    def cancleall():
        theLB.selection_clear(0,END)
    def stdbutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] ==standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                logging.error('Standard is not set!')
                print('KeyError')

    def sambutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] !=standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                print()
            except NameError:
                logging.error('Standard is not set!')
               
            except Exception as e:
                print(str(e))
                logging.error('Unknown erro:%s',str(e))
        

    theButton = Button(frmLB,text='Delete',\
                   command=lambda x=theLB:x.delete(ACTIVE))
  
    theButton.grid(row =11,column =2,pady=5)
    showButton=Button(frmLB,text='All',command=selectall)
    showButton.grid(row =11,column =0,pady=5)
 
    cancleButton=Button(frmLB,text='Cancle',command=cancleall)
    cancleButton.grid(row =11,column =1,pady=5)

   
    return fig_name
    

def instructure2():
    global sampleslist
    global isoname
    sampleslist={}
    os.chdir(outputpath)
    
   
    with open("samples.csv", "a", newline='') as s:
        
        writer=csv.writer(s)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])
            
    with open("std.csv", "a", newline='') as f:
        writer=csv.writer(f)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])
    with open("result_all.csv", "a", newline='') as g:
        writer=csv.writer(g)
        writer.writerow(['No.','FilesName','SamplesName','207Pb/206Pb','2s','206Pb/238U','2s','207Pb/235Uc','2s','208Pb/232Th','2s','208Pb/206Pb','2s','232Th/206Pb','2s','208Pb/204Pb','2s'])      


    samplename=[]
    
    global Numbers
    global Timeinternal
    global fig_name
    global num
    num={}
    root = tk.Tk()
    root.withdraw()
    default_dir=inputpath
    files_list=[]

    for name in os.listdir(inputpath):
        if os.path.splitext(name)[1]=='.FIN':
            file_path=name
            
        elif os.path.splitext(name)[1]=='.FIN2':
            files_list.append(name)
            
    #file_path=tk.filedialog.askopenfilename(title=u'Please Choose the List File', initialdir=(os.path.expanduser(default_dir)))

    listfile=open(default_dir+'//'+file_path)
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    listfile.readline()
    sampleslist={}
    for line in listfile.readlines():
        sampleslist[line[0:-2]+'2']=line.rsplit('-',1)[0]
        
    
    

    num={}

    
        
    window = tk.Tk()
    window.title('Loading......')
    window.geometry('630x150')
    tk.Label(window, text='Progress of loading:', ).place(x=50, y=60)
    canvas = tk.Canvas(window, width=465, height=22, bg="white")
    canvas.place(x=110, y=60)       
    w=Canvas(window, width=200, height=22, bg="pink")              
    w.place(x=250, y=90)  
            
    #print(sampleslist)         

           
        
    i=1
    scale=len(files_list)
    print('Start Loading....'.center(scale//2,'='))
    fill_line = canvas.create_rectangle(1.5, 1.5, 0, 23, width=0, fill="green")
        
    n = 465 / scale  
    for fname in files_list:
        a='*'*i
        b='.'*(scale-i)
        c=(i/scale)*100
        
        
        name=fname
        w.create_text(90,10,text=f'Loading.... {name} ')
        
        isoname=['Time','Hg202','Pb204','Pb206','Pb207','Pb208','Th232','U238']
        x,y1,y2,y3,y4,y5,y6,y7=loaddata(name,isoname)
            
         #Save the signal and background time for automatic identification

        Numbers=len(x)
        Timeinternal=(x[Numbers-1]-x[0])/Numbers

        
        
        b0=bcg_from
        b1=bcg_to
        y8=[]

        bcg_s=int(b0/Timeinternal)
        bcg_e=int(b1/Timeinternal)
        
        for i in range(len(x)):
            sum_a=y1[i]+y2[i]+y3[i]+y4[i]+y5[i]+y6[i]+y7[i]
            y8.append(sum_a)
        signal_bcg=np.array(y8)[bcg_s:bcg_e]
            
            
        signal_bcg=np.where(np.abs(signal_bcg-np.average(signal_bcg))<2*np.std(signal_bcg),signal_bcg,np.average(signal_bcg))   

        bcg=signal_bcg.mean()
    
  
        std=np.std(y8[bcg_s:bcg_e])
        sdmul=multi
        ind = [True if value > bcg+sdmul *std else False for value in y8]
        ind2 = ind[1:]
        ind2.append(False)


        #print(ind)
        #print(ind2)
        index = [i for i in range(0, len(ind)) if ind[i] != ind2[i]]

                
        
        starts = [index[i] for i in range(len(index)) if i % 2 == 0]
        ends = [index[i] for i in range(len(index)) if i % 2 != 0]
        #print(name,':','s',starts,'e',ends)
        if len(ends)>1 and len(starts)>1:
            for i in range(len(starts)):
                posi=0
                if starts[i]<((b1)/Timeinternal):
                    s0=float(starts[posi+1])*abs(Timeinternal)  
                else:
                    s0=float(starts[posi])*abs(Timeinternal)
                pos=posi+1
            for j in range(len(ends)):
                posj=0
                if ends[j]<((b1)/Timeinternal):
                    s1=float(ends[posj+1])*abs(Timeinternal)
                    
                else:
                    s1=float(ends[posj])*abs(Timeinternal)
                posj=posj+1
            

        else:
            s1=float(ends[0])*abs(Timeinternal)
            s0=float(starts[0])*abs(Timeinternal)              

        Diff=(s1-1)-(s0+1)
        if Diff>0:
            num[name]=[Numbers,b0,b1,s0+1,s1-1]
            
        else:num[name]=[Numbers,b0,b1,s1+1,s0-1]          
    
        
        


        np.save(outputpath+'//'+name+'_x',x)
        np.save(outputpath+'//'+name+'_y1',y1)
        np.save(outputpath+'//'+name+'_y2',y2)
        np.save(outputpath+'//'+name+'_y3',y3)
        np.save(outputpath+'//'+name+'_y4',y4)
        np.save(outputpath+'//'+name+'_y5',y5)
        np.save(outputpath+'//'+name+'_y6',y6)
        np.save(outputpath+'//'+name+'_y7',y7)
        
        i=i+1
        n = n + 465 / scale
          
        canvas.coords(fill_line, (0, 0, n, 60))
        window.update()
        time.sleep(0)  
        w.delete("all")
        #print("\r{:^3.0f}%[{}->{}]".format(c,a,b),end='')
    np.save(outputpath+'//'+'Time_setting',num)
    window.destroy()    
    '''if max(num.values())-min(num.values())>=5:
        print()
        tk.messagebox.showinfo(title='Warning！', message='he original data file is abnormal!Plese check：'+str(min(num,key=num.get)))
        logging.critical('Warning notice: The original data file is abnormal!Plese check %s',str(min(num,key=num.get)))

        print('he original data file is abnormal!Plese check：',min(num,key=num.get))
    else:
        '''
    logging.info("Loading compeleted.")   
    print()
    print('Loading compeleted'.center(scale//2,'='))
    print()
    tk.messagebox.showinfo(title='Information！', message='Loading compeleted！Total '+str(scale)+'files.')
    logging.info("Total %s files have been loaded.",str(len(files_list)))  
       
       

    global theLB,sc
    theLB=Listbox(frmLB,selectmode=SINGLE,height=11)
    theLB.grid(row =6,column =0,rowspan=4,columnspan=3,pady=2,padx=5)
    sc = tkinter.Scrollbar(frmLB,command=theLB.yview)
    sc.grid(row =6,column =3,rowspan=4,pady=2,sticky='ns') 

    theLB.bind("<<ListboxSelect>>", showplt)

    theLB.configure(yscrollcommand = sc.set)
    
    for item in sampleslist.keys():
        fig_name=item
        theLB.insert(END,fig_name)
    
    def selectall():
        theLB.select_set(0,END)
        
    def cancleall():
        theLB.selection_clear(0,END)
    def stdbutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] ==standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                print('KeyError')
            except NameError:
                logging.error('Standard is not set!')
                print('Standard is not set！')

    def sambutton():
        theLB.select_set(0,END)
        for i,std in enumerate(theLB.curselection()):
            
            try:
                
                if sampleslist[theLB.get(std)] !=standard:
                    continue
                else:
                    theLB.selection_clear(std)
            except KeyError:
                print()
            except Exception as e:
                print(str(e))
                logging.error('Unknown erro:%s',str(e))
        
        
    theButton = Button(frmLB,text='Delete',command=lambda x=theLB:x.delete(ACTIVE))
  
    theButton.grid(row =11,column =2,pady=5)
    showButton=Button(frmLB,text='All',command=selectall)
    showButton.grid(row =11,column =0,pady=5)
 
    cancleButton=Button(frmLB,text='Cancle',command=cancleall)
    cancleButton.grid(row =11,column =1,pady=5)
    
 
    return fig_name
        

   
        
    
def dataprocess_selected():

    logging.debug('Correction for selected samples.' )
    index = theLB.curselection()
    var=theLB.get(index)
    
    numn=np.load('Time_setting.npy',allow_pickle='True').item()
    b0=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Background Start (second)：',initialvalue = '5')
    
    b1=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Background End (second)：',initialvalue = '22')
    
    s0=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Signal Start (second)：',initialvalue = '30')
    
    s1=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Signal End (second)：',initialvalue = '66')

    
    numn[var][1]=b0
    numn[var][2]=b1
    numn[var][3]=s0
    numn[var][4]=s1

    np.save('Time_setting.npy',numn)
    
    
            
 
                
                

        
    
                    
def settime():
    global start_b
    global end_b
    global start_s
    global end_s
    
    
    
    b0=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Background Start (second)：',initialvalue = '1')
    
    b1=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Background End (second)：',initialvalue = '15')
    
    s0=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Signal Start (second)：',initialvalue = '30')
    
    s1=tk.simpledialog.askfloat(title = 'Intergrating time',prompt='Signal End (second)：',initialvalue = '65')
    
    logging.info('Dwell times:%s;%s;%s;%s ',str(b0),str(b1),str(s0),str(s1))

    num=np.load('Time_setting.npy',allow_pickle='True').item()
    for name in num.keys():
        num[name][1]=b0
        num[name][2]=b1
        num[name][3]=s0
        num[name][4]=s1
        start_b=int(b0/Timeinternal)
        end_b=int(b1/Timeinternal)
        start_s=int(s0/Timeinternal)
        end_s=int(s1/Timeinternal)
    
    np.save('Time_setting.npy',num)
    tk.messagebox.showinfo(title='Information！', message='The integration time is set successfully. Please click [start calculation] to complete the data processing.')
         
        
        
def Sample_Pb_Calculate():
    logging.info("Commom Pb correlation of the samples.")
    Ms=tk.messagebox.askyesno(title = 'Set Pb correction mode！',message="Calculate by the age ('Y') or customize the initial lead isotope  (' N ')?")
    if Ms:
        age=tk.simpledialog.askfloat(title = 'Pb isotopic settings of samples',prompt='Age(Ma):',initialvalue = '100')

        Radioactive_Pb=(1/137.88)*(exp(0.00000000098485*age*1000000)-1)/(exp(0.000000000155125*age*1000000)-1)
        common_Pb=(15.628-(exp(0.00000000098485*age*1000000)-1)*(9.735/137.88))/(18.700-(exp(0.000000000155125*age*1000000)-1)*9.735)
    else:
        Radioactive_Pb=tk.simpledialog.askfloat(title = 'Pb isotopic settings of samples',prompt='Radioactive 207Pb/206Pb:',initialvalue = '0.048015')
        common_Pb=tk.simpledialog.askfloat(title = 'Pb isotopic settings of samples',prompt='Common 207Pb/206Pb:',initialvalue = '0.842185')
    
    logging.info("Common Pb Calculation")
    logging.info("Common Pb :%s",common_Pb)
    logging.info("Radioactive_Pb :%s",Radioactive_Pb)
    try:
        name=np.loadtxt(outputpath+'//'+'result_all.csv',dtype=str,delimiter=',',skiprows=1,usecols=(2))
        ssf=name.tolist()
        date=np.loadtxt(outputpath+'//'+'result_all.csv',delimiter=',',skiprows=1,usecols=(3,4,5,6,7,8,9,10,11,12,13,14,15,16))
        
        date_all=pd.DataFrame(date,index=name).astype(float)
        
        
        
        myix=[i for i,x in enumerate(ssf) if x==standard]
        myiy=[i for i,x in enumerate(ssf) if x!=standard]
        
            
        sample_slected_pd=pd.DataFrame(columns=['0','1','2','3','4','5','6','7','8','9','10','11','12','13'])

        for i in range(len(myix)):
            
            if i==0:
                if myix==0:
                    continue
                else:
                   sample_slected_pd=sample_slected_pd.append(date_all[:][:myix[0]], ignore_index=True) 
            else:
                
                if myix[i]-myix[i-1]==1:
                    continue
                sample_slected_pd=sample_slected_pd.append(date_all[:][myix[i-1]+1:myix[i]], ignore_index=True)
        
                if i==len(myix)-1:
                    sample_slected_pd=sample_slected_pd.append(date_all[:][myix[i]+1:], ignore_index=True)
        
        def func(lst):
            new_list = []
            for i,j in zip(lst,lst[1:]):
                    if j - i > 1:
                            new_list.append(lst[:lst.index(j)])
                            lst = lst[lst.index(j):]
            new_list.append(lst)
            return new_list

        class_sample=func(myiy)
        class_std=func(myix)
        date_all_num=pd.DataFrame(date).astype(float)
        
        
            
        def factor(SEP,stand_v,i,class_std,class_sample):
            
            if len(class_std)==len(class_sample):
                if list(class_std[0])[0]==0:
                    class_std.append(class_std[-1])
                    f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                    return f
                else:
                     class_std.insert(0,class_std[0])
                     f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                     return f
            elif len(class_std)<len(class_sample):
                class_std.append(class_std[-1])
                class_std.insert(0,class_std[0])
                
                f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                return f
            else:
                f=stand_v/date_all_num[i][class_std[SEP][0]:class_std[SEP][-1]+1].mean()              
                return f
            
               
                
        result_cal_age=[]
        try:
            
            
            for SEP,each in enumerate(class_sample):
                
                 
                n=len(each)
                w_1=1
                
                f207_206A=factor(SEP,P382,0,class_std,class_sample)    
                f206_238A=factor(SEP,a,2,class_std,class_sample) 
                f207_235A=factor(SEP,b,4,class_std,class_sample)  
                f208_232A=factor(SEP,c,6,class_std,class_sample)    
                f208_206A=1
                f232_206A=1/f208_232A
                f208_204A=1
                f207_206B=factor(SEP+1,P382,0,class_std,class_sample)    
                f206_238B=factor(SEP+1,a,2,class_std,class_sample) 
                f207_235B=factor(SEP+1,b,4,class_std,class_sample)  
                f208_232B=factor(SEP+1,c,6,class_std,class_sample)    
                f208_206B=1
                f232_206B=1/f208_232B
                f208_204B=1         
               
                for i in each:
                    
                    result_cal_207Pb_206Pb=date_all[0][i]*(w_1/n)*f207_206B+date_all[0][i]*(1-(w_1/n))*f207_206A
            
                    result_cal_206Pb_238U=date_all[2][i]*(w_1/n)*f206_238B+date_all[2][i]*(1-(w_1/n))*f206_238A
                    result_cal_207Pb_235U=date_all[4][i]*(w_1/n)*f207_235B+date_all[4][i]*(1-(w_1/n))*f207_235A
                    result_cal_208Pb_232Th=date_all[6][i]*(w_1/n)*f208_232B+date_all[6][i]*(1-(w_1/n))*f208_232A
                    result_cal_208Pb_206Pb=date_all[8][i]*(w_1/n)*f208_206B+date_all[8][i]*(1-(w_1/n))*f208_206A
                    result_cal_232Th_206Pb=date_all[10][i]*(w_1/n)*f232_206B+date_all[10][i]*(1-(w_1/n))*f232_206A
                    result_cal_208Pb_204Pb=date_all[12][i]*(w_1/n)*f208_204B+date_all[12][i]*(1-(w_1/n))*f208_204A
                    
                    w=(sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U)/result_cal_206Pb_238U*100
                    s=(sqrt(pow((date_all[5][i]/result_cal_207Pb_235U),2)+0.0009)*result_cal_207Pb_235U)/result_cal_207Pb_235U*100
                    o=(sqrt(pow((date_all[1][i]/result_cal_207Pb_206Pb),2)+0.0001)*result_cal_207Pb_206Pb)/result_cal_207Pb_206Pb*100
                    if (abs((w*w+s*s-o*o)/(2*w*s)))<w/s:
                        rho=abs((w*w+s*s-o*o)/(2*w*s))
                    else :
                        rho=w/s
                        
                    s002=[ssf[i],result_cal_207Pb_206Pb,date_all[1][i],result_cal_206Pb_238U,date_all[3][i],result_cal_207Pb_235U,date_all[5][i],result_cal_208Pb_232Th,date_all[7][i],\
                          result_cal_208Pb_206Pb,date_all[9][i],
                          result_cal_232Th_206Pb,date_all[11][i],result_cal_208Pb_204Pb,date_all[13][i],
                          '------------',log(abs(result_cal_206Pb_238U+1))/(0.000000000155125)/1000000,(sqrt(pow(date_all[3][i]/result_cal_206Pb_238U,2)+0.0009))*(log(abs(result_cal_206Pb_238U+1))/(0.000000000155125)/1000000),\
                          log(abs(result_cal_207Pb_235U+1))/(0.00000000098485)/1000000,(sqrt(pow(date_all[5][i]/result_cal_207Pb_235U,2)+0.0009))*(log(abs(result_cal_207Pb_235U+1))/(0.00000000098485)/1000000),\
                          log(abs(result_cal_208Pb_232Th+1))/(0.000000000049475)/1000000,(sqrt(pow(date_all[7][i]/result_cal_208Pb_232Th,2)+0.0009))*(log(abs(result_cal_208Pb_232Th+1))/(0.000000000049475)/1000000), Age76Pb(result_cal_207Pb_206Pb),Age76Pb(result_cal_207Pb_206Pb+date_all[1][i])-Age76Pb(result_cal_207Pb_206Pb),'--------',\
                          (result_cal_207Pb_206Pb-Radioactive_Pb)/(common_Pb-Radioactive_Pb),\
                          (log(abs((1-(result_cal_207Pb_206Pb-Radioactive_Pb)/(common_Pb-Radioactive_Pb))*result_cal_206Pb_238U)+1)/0.000000000155125)/1000000,\
                          ((log(abs((1-(result_cal_207Pb_206Pb-Radioactive_Pb)/(common_Pb-Radioactive_Pb))*result_cal_206Pb_238U)+1)/0.000000000155125)/1000000)*\
                          ((sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U)/result_cal_206Pb_238U*100)/100,\
                          '------------',1/result_cal_206Pb_238U,(1/result_cal_206Pb_238U)*(sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U)/result_cal_206Pb_238U,\
                          result_cal_207Pb_206Pb,date_all[1][i],'------------',result_cal_207Pb_235U,result_cal_207Pb_235U*sqrt(pow(date_all[5][i]/result_cal_207Pb_235U,2)+0.0009),result_cal_206Pb_238U,sqrt(pow((date_all[3][i]/result_cal_206Pb_238U),2)+0.0009)*result_cal_206Pb_238U,\
                          rho,                       
                             
                          result_cal_208Pb_232Th,date_all[7][i],result_cal_208Pb_206Pb,date_all[9][i]
                          ]
                    
                    
                    result_cal_age.append(s002)
                    w_1=w_1+1
        except Exception as e:
            print(str(e))
            logging.error('Unknown erro:%s',str(e))
            tk.messagebox.showinfo(title='Information！', message='Standard  or “result_all”file erros ！')
           
                
                
                
        

            
        
        os.chdir(outputpath)

        workbook = xlwt.Workbook()
        worksheet=workbook.add_sheet('My Worksheet',cell_overwrite_ok=True)
        style = xlwt.XFStyle() 
        font = xlwt.Font() 
        font.name = 'Times New Roman' 
        font.bold = True
        font.underline = True 
        font.italic = True 
        style.font = font     
        worksheet.write(0, 0, 'SampleName')
        worksheet.write(0, 1, '207Pb/206Pb')
        worksheet.write(0, 2, '2s')
        worksheet.write(0, 3, '206Pb/238U')
        worksheet.write(0, 4, '2s')

        worksheet.write(0, 5, '207Pb/235U')
        worksheet.write(0, 6, '2s')

        worksheet.write(0, 7, '208Pb/232Th')
        worksheet.write(0, 8, '2s')
        worksheet.write(0, 9, '208Pb/206Pb')
        worksheet.write(0, 10, '2s')
        
        worksheet.write(0, 11, '232Th/206Pb')
        worksheet.write(0, 12, '2s')
        worksheet.write(0, 13, '208Pb/204Pb')
        worksheet.write(0, 14, '2s')
        
        worksheet.write(0, 15, 'Age')
        worksheet.write(0, 16, '206Pb_238U Age(Ma)')
        worksheet.write(0, 17, '2s')
        worksheet.write(0, 18, '207Pb_235U Age(Ma)')
        worksheet.write(0, 19, '2s')    
        worksheet.write(0, 20, '208Pb_232Th  Age(Ma)')
        worksheet.write(0, 21, '2s')
        worksheet.write(0, 22, '207Pb_206Pb  Age(Ma)')
        worksheet.write(0, 23, '2s')
        
        worksheet.write(0, 24, 'Corrected 206Pb_238U Age')
        worksheet.write(0, 25, 'f206')
        worksheet.write(0, 26, 'Rad 206Pb_238U Age(Ma)')
        worksheet.write(0, 27, '2s')

        worksheet.write(0, 28, 'Terra-Wasserburg Plot:')
        worksheet.write(0, 29, '238U/206Pb')
        worksheet.write(0, 30, '2s error')
        worksheet.write(0, 31, '207Pb/206Pb')
        worksheet.write(0, 32, '2s error')
        worksheet.write(0, 33, 'Plotting purposes:')
        worksheet.write(0, 34, '207Pb/235U')
        worksheet.write(0, 35, '2s error')
        worksheet.write(0, 36, '206Pb/238U')
        worksheet.write(0, 37, '2s error')
        worksheet.write(0, 38, 'rho')
        worksheet.write(0, 39, '208Pb/232Th')
        worksheet.write(0, 40, '2s error')
        worksheet.write(0, 41, '208Pb/206Pb')
        worksheet.write(0, 42, '2s error')
        
        
    
        
        
       
        
        
        for row in range(len(result_cal_age)):
            for col in range(43):
                worksheet.write(row+1,col,result_cal_age[row][col])
            
            
        
        now = time.strftime("%H%M%S",time.localtime(time.time()))     
        workbook.save('Age_result_corPb'+now+'.xls')
        
        logging.info('The calculation was successful.The file saved in:%s',outputpath)
        
    except ValueError:
        
        tk.messagebox.showinfo(title='Information！', message='“result_all.csv" file not found or  Incorrect data format！')
        logging.critical('“result_all.csv" were not found or the format is incorrect.')  
    except NameError:
        
       
        tk.messagebox.showinfo(title='Information！', message='Output directory or standard  is not set！')
        logging.error('The output directory is not set.') 
    except OSError:
        tk.messagebox.showinfo(title='Information！', message='"result_all.csv" not found！')
        logging.error('"result_all.csv" not found')
        print('"result_all.csv" not found')
        
    tk.messagebox.showinfo(title='Congratulations！', message='Common Pb correction has been completed！')
            

def main():
    #sys.stdout = Logger(sys.stdout)
    #sys.stderr = Logger(sys.stderr)
    global Numbers
    logging.info("Program Start.")
    global bcg_from
    global bcg_to
    global multi
    global var0
    global var1
    global var3
    global var4
    global num
    global start_b
    global start_s
    global end_s
    global end_b
    multi=8

    global bcg
    bcg=[1,5,8]
    filename = open('bcg_inf.txt', 'w')
    for value in bcg:
        filename.write(str(value))
    filename.close()
        
         
    
    inputpath = ''
    outputpath=''
    global regression_method
    regression_method=1


    def selectPath_in():
        global inputpath        
        inputpath=askdirectory(title=u'Select input folder:',initialdir=(os.path.expanduser('H:/')))
        print('Open：', inputpath)
        if inputpath is not None:
            path1.set(inputpath)
           
        
    
    def selectPath_out():
        global outputpath
        global theLB
        outputpath=askdirectory(title=u'Select output folder',initialdir=(os.path.expanduser('H:/')))
        print('Save：',outputpath)
        
        if outputpath is not None:
            path2.set(outputpath)
            




             

    def software():
        tk.messagebox.showinfo(title='About software！', message='Version 1.8\nShenzhen Jianshe Tong Engineering Technology Co., Ltd. \nCopyright (C) 2021   All Rights Reserved.')
             
    def developer():
        tk.messagebox.showinfo(title='About author！', message='Author：Guoqi Liu  Email: 642847452@qq.com  Tel:18942230808')
    
    
    t=Tk()
    
    t.title('Isoclock v1.8 ' )
    menubar = Menu(t)
    t.config(menu=menubar)
    helpmenu = Menu(menubar, tearoff=0)
    menubar.add_cascade(label="About", menu=helpmenu)
    #helpmenu.add_command(label="Instructions", command="")
    helpmenu.add_command(label="About software", command=software)
    helpmenu.add_command(label="About developer", command=developer)
    

    
    frmTT=Frame(width=810, height=50,bg='#B0C4DE')
    frmLT = Frame(width=810, height=50,bg='#008080')
    frmLLT = Frame(width=810, height=50,bg='#D2691E')
    frmLC = Frame(width=405, height=100,bg='#FDDA86')
    global frmLB
    frmLB = Frame(width=405, height=330,bg='#CCA5E2')
    
    global frmRB
    frmRB = Frame(width=405, height=330)
    
    frmRC = Frame(width=405, height=50,bg='#6493AF')
     
    
    frmRT = Frame(width=405, height=50,bg='#9ACD32')
    


    def Cal_age(age):
        global R8,Q8,S8,P382,a,b,c
        a=exp(0.000000000155125*age*1000000)-1
        b=exp(0.00000000098485*age*1000000)-1
        c=exp(0.000000000049475*age*1000000)-1
        P382=1/137.88*(b/a)
            
        Q8=18.700-a*9.735
        R8=15.628-b*(9.735/137.88)            
        S8=38.630-c*36.837
        return R8,Q8,S8,P382,a,b,c


    
    global Standard_names
    
    Standard_names={}
    Standard_names['AY-4']=158
    Standard_names['HK-30']=151
    
    Standard_names['MAD-1']=485
    Standard_names['MAD-2']=485
    Standard_names['MAD-3']=485
    Standard_names['MAD-NEW']=485
    Standard_names['YGX2113']=160
    Standard_names['BLR']=1047
    Standard_names['OLT']=1007
    Standard_names['Ont']=1054
    
    
    def userset_standard():
       
        M1=tk.messagebox.askyesno(title = 'Standard Setting Mode！',message="Calculated by Age('Y') or entry Initial Pb isotope('N')?")
        global R8,Q8,S8,P382,a,b,c
        if M1:
            
            Standard_age=tk.simpledialog.askfloat(title = 'Age of Standard ',prompt='Age of Standard（Ma）',initialvalue = '485')           
            Standard_name=tk.simpledialog.askstring(title = 'Name of Standard',prompt='Name of Standard',initialvalue = 'MAD')
            
            Standard_names[Standard_name]=Standard_age
            
            R8,Q8,S8,P382,a,b,c=Cal_age(Standard_age)
            logging.info(Standard_names)
                  
            logging.info('207Pb/204Pb%s',str(R8))
            logging.info('206Pb/204Pb:%s',str(Q8))
            logging.info('208Pb/204Pb:%s',str(S8))
            logging.info('207Pb/206Pb:%s',str(P382))
            logging.info('206Pb/238U:%s',str(a))
            logging.info('207Pb/235U:%s',str(b))
            logging.info('208Pb/232Th:%s',str(c))
            
        else :
            
            Standard_name=tk.simpledialog.askstring(title = 'Name of Standard',prompt='Name of Standard',initialvalue = 'MAD')
            Standard_age=tk.simpledialog.askfloat(title = 'Age of Standard',prompt='Age of Standard（Ma）',initialvalue = '485')
            Standard_names[Standard_name]=Standard_age
            R8,Q8,S8,P38233,a33,b,c=Cal_age(Standard_age)
            #R8=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='207Pb/204Pb:',initialvalue = '15.616')
            #Q8=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='206Pb/204Pb:',initialvalue = '18.458')
            #S8=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='208Pb/204Pb:',initialvalue = '38.341')
            P382=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='207Pb/206Pb:',initialvalue = '0.049213')
            a=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='206Pb/238U:',initialvalue = '0.024813')
            #b=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='207Pb/235U:',initialvalue = '0.168366')
            #c=tk.simpledialog.askfloat(title = 'Setting standard isotope',prompt='208Pb/232Th:',initialvalue = '0.007848')
            logging.info(Standard_names)
            
            logging.info('207Pb/204Pb:%s',str(R8))
            logging.info('206Pb/204Pb:%s',str(Q8))
            logging.info('208Pb/204Pb:%s',str(S8))
            logging.info('207Pb/206Pb:%s',str(P382))
            logging.info('206Pb/238U:%s',str(a))
            logging.info('207Pb/235U:%s',str(b))
            logging.info('208Pb/232Th:%s',str(c))
            
        return Standard_names,Standard_name
    def choose_result1():              
        global standard
        Cal_age(154)
        
        standard='AY-4'
        logging.info('Choose Cassiterite U-Pb Dating，use <AY-4> as standard!')    
        return standard     
    def choose_result2():        
        global standard
        Cal_age(483)
        
        standard='MAD-NEW'
        
     
    
         
        logging.info('Choose Apatite U-Pb Dating，use <MAD> as standard!')           
        return standard
    def choose_result3():
        global standard
        standard='91500'
        Cal_age(1064.5)
       
        logging.info('Choose Titanite U-Pb Dating，use <91500> as standard !')
        return standard
                   
    def choose_result4():
        global standard
        Cal_age(254.4)
        standard='wc-1' 

          
        logging.info('Choose Wolframite U-Pb Dating，use <wc-1> as standard!')               
        return standard
    def choose_result5():
        global standard
        
        Standard_names,standard=userset_standard()
        
        logging.info(Standard_names)        
        logging.info('Standard %s has been set!',standard) 
        return Standard_names,standard

    def setbcg_multi():
        numn=np.load('Time_setting.npy',allow_pickle='True').item()
        filename = open('bcg_inf.txt', 'w')
        
        bcg_from=float(var0.get())
        bcg_to=float(var1.get())
        
        sig_from=float(var3.get())
        sig_to=float(var4.get())
        for name in numn.keys():
            numn[name][1]=bcg_from
            numn[name][2]=bcg_to
            numn[name][3]=sig_from
            numn[name][4]=sig_to
        np.save('Time_setting.npy',numn)
        bcg=[bcg_from,bcg_to,8]
        filename = open('bcg_inf.txt', 'w')
        for value in bcg:
            filename.write(str(value))
        filename.close()
    
    def setbcg_single():
            index = theLB.curselection()
            var=theLB.get(index)
    
            numn=np.load('Time_setting.npy',allow_pickle='True').item()
            b0=float(var0.get())
    
            b1=float(var1.get())
    
            s0=float(var3.get())
    
            s1=float(var4.get())

    
            numn[var][1]=b0
            numn[var][2]=b1
            numn[var][3]=s0
            numn[var][4]=s1

            np.save('Time_setting.npy',numn)
            showplt(theLB.curselection())

        
    
    
    v=IntVar()
    v.set(3)
    global stdcor
    stdcor=IntVar()
    stdcor.set(0)
    
    stdcor_method=Radiobutton(frmTT, text="Standard Cor", width = 15,variable=stdcor,value=0)
    std_method=Radiobutton(frmTT, text="Standard not Cor ", width = 15,variable=stdcor,value=1)
    
    global theLB,sc
    theLB=Listbox(frmLB,selectmode=SINGLE,height=11)
    theLB.grid(row =6,column =0,rowspan=4,columnspan=3,pady=2,padx=5)
    sc = tkinter.Scrollbar(frmLB,command=theLB.yview)
    sc.grid(row =6,column =3,rowspan=4,pady=2,sticky='ns')
   
    
    age=Button(frmRC,text="Calculation (linear)",command=Age_Calculate)
    age2=Button(frmRC,text="Calculation (average)",command=Age_Calculate_average)
    Pb_deduction=Button(frmRC,text="Pb correction",command=Sample_Pb_Calculate)
    user_methodbutton=Radiobutton(frmLT, text="Setting Standard", width = 15,variable=v,value=5,command=choose_result5)

    choose_method1button=Radiobutton(frmLT, text="Cassiterite(AY-4)",width = 13,variable=v,value=1,command=choose_result1)
    choose_method2button=Radiobutton(frmLT, text="Apatite(MAD-NEW)", width = 15,variable=v,value=2,command=choose_result2)
    choose_method3button=Radiobutton(frmLT, text="Zircon(91500)", width =10,variable=v,value=3,command=choose_result3)
    choose_method4button=Radiobutton(frmLT, text="Calcite(wc-1)", width =15,variable=v,value=4,command=choose_result4)


        
 
 
   
    
    path1=StringVar()
    sr1=Label(frmLC,text = "Input Path:")

    sr2=Entry(frmLC, textvariable = path1)
    
    sr3=Button(frmLC, text = "Browse", command = selectPath_in)
    

    path2=StringVar()
    sc1=Label(frmLC,text = "Output Path:")
    sc2=Entry(frmLC, textvariable =path2)
    sc3=Button(frmLC, text = "Browse", command = selectPath_out)
    def samplelist():
        logging.info('start loading date.')
        if ele.get()==0:
            logging.info('Choose "Thermo" instrument.')
            instructure0()
        elif ele.get()==1 :
            logging.info('Choose "Agilent" instrument.')
            instructure1()
        elif ele.get()==2 :
            logging.info('Choose "Elenment" instrument.')
            instructure2()
        
    
    l=Label(frmRB)
    read_button=Button(frmLC, text = "Load\nData", command=samplelist)
    
    global ele   
    ele=IntVar()
    ele.set(0)
    instrument0=Radiobutton(frmLLT, text="Thermo",width = 8,variable=ele,value=0)
    instrument1=Radiobutton(frmLLT, text="Agilent",width = 8,variable=ele,value=1)
    instrument2=Radiobutton(frmLLT, text="Element",width = 8,variable=ele,value=2)

    
    

            
    
    plot_button=Button (frmLB, text = "Samples List",width=30)
   
   
            
    var_s=IntVar()
    cal_button=Button(frmRT, text = "Pb correction (All)", command = dataprocess)    

    
    set_time_all = Button(frmLB, text='Signal selection',width=20)

    #user set  background and sample during time.

    
    #Set background and  multiply
    def test(content):
        if content.isdigit() or content=='':
            return True
        else:
            return False
 
    var0 = StringVar()
    var1 = StringVar()
    var3 = StringVar()
    var4 = StringVar()
    
    var0.set(3)
    var1.set(10)
    var3.set(30)
    var4.set(60)
    
    bcg_from=float(var0.get())
    bcg_to=float(var1.get())
    sig_from=float(var3.get())
    sig_to=float(var4.get())
    
    entry0 =Entry(frmLB, bd=0,bg='lightcyan',width =6,textvariable=var0,validate='key',validatecommand=(test,'%P'))
    entry0.grid(row =6,column =6)
    entry1=Entry(frmLB, bd=0,bg='lightcyan',width =6,textvariable=var1,validate='key',validatecommand=(test,'%P'))
    entry1.grid(row =7,column =6)
    entry_3=Entry(frmLB, bd=0,bg='lightcyan',width =6,textvariable=var3,validate='key',validatecommand=(test,'%P'))
    entry_3.grid(row =8,column =6)
    entry_4=Entry(frmLB, bd=0,bg='lightcyan',width =6,textvariable=var4,validate='key',validatecommand=(test,'%P'))
    entry_4.grid(row =9,column =6)
    
    Label0=Label(frmLB,width=8,text='Bcg start(s):',padx=5)
    Label0.grid(row =6,column =5)
    Label1=Label(frmLB,width=8,text='Bcg end(s):',padx=5)
    Label1.grid(row =7,column =5)
    Label_3=Label(frmLB,width=8,text='Sig start(s):',padx=5)
    Label_3.grid(row =8,column =5)
    Label_4=Label(frmLB,width=8,text='Sig end(s):',padx=5)
    Label_4.grid(row =9,column =5)
    

    
    ApplyButton=Button(frmLB,text='Set (All)',command=setbcg_multi,width=10)
    ApplyButton.grid(row =10,column =6)
    ClearButton=Button(frmLB,text='Set (Single)',command=setbcg_single,width=10)
    ClearButton.grid(row =10,column =5)
    
    frmTT.grid(row=0,column=0,columnspan=8,rowspan=1)
    frmLT.grid(row=1,column=0,columnspan=8,rowspan=1)
    frmLLT.grid(row=2,column=0,columnspan=8) 
    frmRT.grid(row = 3,column = 5,columnspan=2)
     
    frmLC.grid(row =3,column = 0,rowspan=2)
    frmRC.grid(row=4,column = 5,columnspan=2)
    frmLB.grid(row =5,column = 0,rowspan=3)
    frmRB.grid(row =5,column = 5,rowspan=3,columnspan=2)
    
    
    


    frmTT.grid_propagate(0)
    frmLT.grid_propagate(0)
    frmLLT.grid_propagate(0)
    frmRT.grid_propagate(0)
    frmLC.grid_propagate(0)
    frmLB.grid_propagate(0)
    frmRC.grid_propagate(0)
    frmRB.grid_propagate(0)
    
    stdcor_method.grid(row=0,column=0,padx=5,pady=10)
    std_method.grid(row=0,column=1,padx=5,pady=10)
    instrument0.grid(row=2,column=0,pady=10,padx=5)
    instrument1.grid(row=2,column=1,pady=10,padx=5)
    instrument2.grid(row=2,column=2,pady=10,padx=5)
    user_methodbutton.grid(row=1,column=0,padx=5,pady=10)

   
    Pb_deduction.grid(row=4,column=7,pady=10,padx=10)





    
    choose_method1button.grid(row=1,column=2,pady=10,padx=5)
    choose_method2button.grid(row=1,column=1,pady=12,padx=5)
    choose_method3button.grid(row=1,column=3,pady=10,padx=5)
    choose_method4button.grid(row=1,column=4,pady=12,padx=5)
    
    cal_button.grid(row =3,column = 5,padx=10,pady=10)
    
    sr1.grid(row =3,column =0,pady=10,padx=5)
    sr2.grid(row =3,column = 1,columnspan=3,pady=10,padx=5)
    sr3.grid(row =3,column =5,pady=10,padx=5)
    sc1.grid(row =4,column = 0,pady=10,padx=5)
    sc2.grid(row =4,column = 1,columnspan=3,pady=10,padx=5)
    sc3.grid(row =4,column = 5,pady=10,padx=5)
    read_button.grid(row=3,column = 6,rowspan=2,pady=15,padx=25)
    
    age.grid(row=4,column =5,pady=10,padx=10)
    age2.grid(row=4,column =6,pady=10,padx=10)
    
    l.grid(row =5, column =5,padx=5,pady=10)
    plot_button.grid(row =5,column = 0,columnspan=3,pady=10,padx=5)
   
    set_time_all.grid(row =5, column =3,columnspan=4,padx=5,pady=10)
   
    def on_closing():
        if tk.messagebox.askokcancel("Quit", "Do you want to quit?"):
            
            t.destroy()
            logging.info("Program Exit.")
    t.protocol("WM_DELETE_WINDOW", on_closing)
                  



    t.mainloop()
    
if __name__=='__main__':
  
    main()   
